#!/usr/bin/env python3
"""
Import P1 field sheet (spring 2025 planting data) into farmOS.

Parses the horizontal multi-section P1 spreadsheet format and creates:
1. Plant assets — named "{planted_date} - {farmos_name} - {section_id}"
2. Quantity entities — inventory counts (measure=count, adjustment=reset)
3. Transplanting logs — backdated to planting date (is_movement=true)

P1 spreadsheet format:
  - Horizontal layout: each section = 4 columns (what, qty, when, p/s)
  - R1 tab: section IDs in Row 1, headers Row 3, data from Row 4
  - R3/R5 tabs: section IDs in Row 2, headers Row 4, data from Row 5
  - P/S column: p=plant (transplant), s=seed (seeding)
  - Both plants and seeds create Plant assets with transplanting logs

Handles edge cases:
  - Compact format in P1R1.29-39: "13 tomato-p" qty=date
  - European decimals: "0,5"
  - Text quantities: "5m (225 seeds)", "1 bag", "10 seeds"
  - Composite species: "Basil sweet+Thai", "Tomatoes Oxheart+Vince"
  - French instructions in cells (filtered as noise)

Usage:
    python scripts/import_p1.py --dry-run                    # Preview all
    python scripts/import_p1.py --row P1R1 --dry-run         # Preview R1 only
    python scripts/import_p1.py                              # Live import
    python scripts/import_p1.py --xlsx path/to/file.xlsx     # Custom file
"""

import argparse
import csv
import os
import re
import sys
import urllib.parse
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

try:
    from farmOS import farmOS as farmOS_client
except ImportError:
    print("ERROR: farmOS library not installed!")
    print("Please install it with: pip install farmOS")
    sys.exit(1)


PLANT_UNIT_UUID = "2371b79e-a87b-4152-b6e4-ea6a9ed37fd0"
AEST = timezone(timedelta(hours=10))

DEFAULT_XLSX = "fieldsheets/P1.2025.Spring Seeding Transplanting.xlsx"
ALT_XLSX = os.path.expanduser(
    "~/Downloads/P1.2025.Spring Seeding Transplanting.xlsx"
)
PLANT_TYPES_CSV = "knowledge/plant_types.csv"

# ── Species name mapping ────────────────────────────────────

# Maps raw P1 spreadsheet names → farmos_name (v8 taxonomy)
SPECIES_MAPPING = {
    # Cabbage variants
    'cabbage "golden acre"': "Cabbage (Golden Acre)",
    'cabbage "red"': "Cabbage (Red)",
    'cabbage "pak choi"': "Pak Choi",
    # Chilli variants
    "chilli-birdeye": "Chilli",
    "chilli-jalapenos": "Chilli (Jalapeño)",
    'chilli "jalapeño"': "Chilli (Jalapeño)",
    "jalapeno": "Chilli (Jalapeño)",
    # Tomato variants
    "vince tomato": "Tomato",
    'tomatoes "mortgage lifter"': "Tomato",
    'tomatoes "mortgage lifter"?': "Tomato",
    'tomatoes "oxheart"+"vince 2 12-13"': "Tomato",
    'tomatoes "rouge de marmande"+"money maker"': "Tomato",
    'tomatoes "vince" + "mortgage lifter"': "Tomato",
    "tomato": "Tomato",
    # Bean variants
    'bean climbing "blue lake"': "Bean - Climbing (Blue Lake)",
    'bean dwarf "borlotti"': "Bean - Dwarf (Borlotti)",
    'bean dwarf "brown beauty"': "Bean - Dwarf",
    'bean dwarf "hawskesbury"': "Bean - Dwarf",
    'bean dwarf "purple beauty"': "Bean - Dwarf",
    "bean dwarf borlotti": "Bean - Dwarf (Borlotti)",
    "beans": "Bean - Dwarf",
    "ffc beans": "Bean - Dwarf",
    "2025.ffc mix beans (borlotti, climbing bean)": "Bean - Dwarf (Borlotti)",
    # Eggplant
    'eggplant "long purple"': "Eggplant (Long Purple)",
    'eggplant "long purple"/2 abbo 12-13': "Eggplant (Long Purple)",
    # Cucurbits
    'cucumber "burpless"': "Cucumber (Burpless)",
    'zucchini "black jack"': "Zucchini (Blackjack)",
    "zucchini lebanese": "Zucchini (Lebanese)",
    "zucchini": "Zucchini (Lebanese)",
    'butternut "ffc"': "Butternut Pumpkin",
    "butternut ffc": "Butternut Pumpkin",
    "pumpkin": "Pumpkin (Queensland Blue)",
    # Alliums
    'onion "long tropea red"': "Onion (Red)",
    "onions and chives": "Onion",
    "spring onion": "Spring Onion",
    # Carrots
    'carrot "baby amsterdam"': "Carrot",
    'carrot "all year round"': "Carrot",
    # Herbs
    "basil sweet": "Basil - Sweet (Classic)",
    "basil thai": "Basil - Perennial (Thai)",
    'basil "sweet"+"thai"': "Basil - Sweet (Classic)",
    "coriander ffc": "Coriander",
    "coriander": "Coriander",
    "rosemary cutting": "Rosemary",
    "mint cutting": "Mint",
    "lavender cutting": "Lavender",
    "geranium cutting": "Geranium",
    "sage": "Sage",
    "fennel ffc": "Fennel",
    "celery": "Celery",
    # Capsicums
    "red capsicum": "Capsicum (Red)",
    # Grains and green manure
    "corn manning pride": "Corn (Manning Pride)",
    "corn": "Corn",
    "ffc corn": "Corn",
    "millet white french": "Millet (White French)",
    "millet gp": "Millet",
    "millet": "Millet",
    "buckwheat ffc": "Buckwheat",
    "buckwheat": "Buckwheat",
    "oat": "Oat",
    "barley": "Barley",
    # Legumes (green manure)
    "cowpea": "Cowpea (Red)",
    "cowpea innoculated": "Cowpea (Red)",
    "sun hemp": "Sunn Hemp",
    "sunn hemp": "Sunn Hemp",
    "sunn hemp+ cowpea": "Sunn Hemp",
    "mix cowpea + sunnhemp (barely-innoculated)": "Cowpea (Red)",
    "mung bean": "Mung Bean",
    "pea bush sugar snap bon": "Pea-Sugar Snap",
    "pea bush sugar snap bon innoculated": "Pea-Sugar Snap",
    "clover crimson": "Clover (Crimson)",
    "red clover": "Clover (Red Persian)",
    "clover red persian clover innoculated": "Clover (Red Persian)",
    "clover white innoculated": "Clover (White Haifa)",
    # Trees and perennials
    "comfrey (p1r2)": "Comfrey",
    "comfrey": "Comfrey",
    "ice cream bean": "Ice Cream Bean",
    "pigeon pea": "Pigeon Pea",
    "tree lucerne - tagasaste": "Tagasaste - Tree Lucerne",
    "citrus seedling": "Citrus (Yuzu)",
    # Others
    "broccoli": "Broccoli (Summer Green)",
    "calendula": "Calendula",
    "marigold": "Marigold (French - Dwarf)",
    "nasturtium": "Nasturtium",
    "okra \"royal burgundi\"": "Okra",
    "parsnip": "Parsnip (Hollow Grown)",
    'spinach "perpetual"': "Spinach-Perpetual",
    'radish "white-tillage"': "Radish (Tillage)",
    "watermelon": "Watermelon (Sugar baby)",
}

# Compact format: "13 tomato-p" → (count, species, type)
COMPACT_SPECIES = {
    "tomato": "Tomato",
    "mint": "Mint",
    "jalapeno": "Chilli (Jalapeño)",
    "sage": "Sage",
    "celery": "Celery",
    "pigeon pea": "Pigeon Pea",
    "beans": "Bean - Dwarf",
    "corn": "Corn",
    "nasturtium": "Nasturtium",
    "oat": "Oat",
    "cowpea": "Cowpea (Red)",
    "sunn hemp": "Sunn Hemp",
}


def load_plant_types():
    """Load plant types CSV for strata lookup."""
    db = {}
    csv_path = Path(PLANT_TYPES_CSV)
    if not csv_path.exists():
        return db
    with open(csv_path) as f:
        for row in csv.DictReader(f):
            fn = row["farmos_name"].strip()
            db[fn] = {
                "strata": row.get("strata", "").strip(),
                "succession_stage": row.get("succession_stage", "").strip(),
            }
    return db


def normalize_species(raw_name):
    """Normalize a P1 species name to farmos_name."""
    if not raw_name:
        return None

    cleaned = raw_name.strip()
    lower = cleaned.lower()

    # Direct mapping
    if lower in SPECIES_MAPPING:
        return SPECIES_MAPPING[lower]

    # Strip suffixes: ffc, cutting, seedling, innoculated
    stripped = re.sub(r"\s+(ffc|cutting|seedling|innoculated)\s*$", "", lower, flags=re.I).strip()
    if stripped in SPECIES_MAPPING:
        return SPECIES_MAPPING[stripped]

    return None


def parse_quantity(raw_qty):
    """Parse quantity from various P1 formats. Returns (count, unit).
    unit is 'plants' for plant counts or 'grams' for seed weights.
    """
    if raw_qty is None:
        return None, None

    s = str(raw_qty).strip()
    if not s:
        return None, None

    # European decimal: "0,5" → 0.5
    if re.match(r"^\d+,\d+$", s):
        s = s.replace(",", ".")

    # Pure number
    try:
        val = float(s)
        return int(val) if val == int(val) else val, "plants"
    except ValueError:
        pass

    # "5m (225 seeds)" → 225
    m = re.search(r"(\d+)\s*seeds?", s, re.I)
    if m:
        return int(m.group(1)), "seeds"

    # "1 bag" → 1
    m = re.match(r"^(\d+)\s*bag", s, re.I)
    if m:
        return int(m.group(1)), "bags"

    # "4/10 and sept" → just take first number
    m = re.match(r"^(\d+)", s)
    if m:
        return int(m.group(1)), "plants"

    return None, None


def parse_compact_entry(text):
    """Parse compact format like '13 tomato-p' or 'ffc beans-s'.
    Returns (count, farmos_name, entry_type) or None.
    """
    text = text.strip()

    # Pattern: "N species-p/s"
    m = re.match(r"^(\d+)\s+(.+?)[-\s]*(p|s)$", text, re.I)
    if m:
        count = int(m.group(1))
        species_raw = m.group(2).strip().lower()
        entry_type = m.group(3).lower()

        # Look up species
        if species_raw in COMPACT_SPECIES:
            return count, COMPACT_SPECIES[species_raw], entry_type
        if species_raw in SPECIES_MAPPING:
            return count, SPECIES_MAPPING[species_raw], entry_type
        return count, None, entry_type

    # Pattern: "species-p/s" (no count, like "ffc beans-s", "nasturtium.s", "oat-s")
    m = re.match(r"^(.+?)[.\-\s]+(p|s)$", text, re.I)
    if m:
        species_raw = m.group(1).strip().lower()
        entry_type = m.group(2).lower()

        if species_raw in COMPACT_SPECIES:
            return None, COMPACT_SPECIES[species_raw], entry_type
        if species_raw in SPECIES_MAPPING:
            return None, SPECIES_MAPPING[species_raw], entry_type
        # Try stripping common prefixes
        stripped = re.sub(r"^(ffc|2025\.ffc)\s+", "", species_raw).strip()
        if stripped in COMPACT_SPECIES:
            return None, COMPACT_SPECIES[stripped], entry_type
        if stripped in SPECIES_MAPPING:
            return None, SPECIES_MAPPING[stripped], entry_type
        return None, None, entry_type

    return None


def parse_date_value(raw_date):
    """Parse date from Excel cell. Returns ISO date string or None."""
    if raw_date is None:
        return None

    if isinstance(raw_date, datetime):
        return raw_date.strftime("%Y-%m-%d")

    s = str(raw_date).strip()
    if not s:
        return None

    # Try ISO format
    try:
        dt = datetime.strptime(s[:10], "%Y-%m-%d")
        return dt.strftime("%Y-%m-%d")
    except ValueError:
        pass

    # "4/10 and sept" → try to parse
    m = re.match(r"^(\d{1,2})/(\d{1,2})", s)
    if m:
        day, month = int(m.group(1)), int(m.group(2))
        if 1 <= month <= 12 and 1 <= day <= 31:
            return f"2025-{month:02d}-{day:02d}"

    return None


def is_noise(text):
    """Check if text is noise (French instructions, headers, etc.)."""
    lower = text.strip().lower()
    if lower.startswith(("tous les", "entre les", "a cote", "au milieu")):
        return True
    if lower in ("what", "species", "total", "notes", "", "portion name",
                  "existing", "plant mix"):
        return True
    return False


# ── Spreadsheet parser ───────────────────────────────────────


def parse_p1_spreadsheet(xlsx_path):
    """Parse P1 spreadsheet into a list of section entries.

    Returns dict: section_id → list of {species, count, date, entry_type, notes}
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    all_sections = {}

    tab_configs = [
        ("P1R1-spring 2025", 1, 4),   # section IDs in Row 1, data from Row 4
        ("P1R3-spring 2025", 2, 5),   # section IDs in Row 2, data from Row 5
        ("P1R5-Spring 2025", 2, 5),   # section IDs in Row 2, data from Row 5
    ]

    for tab_name, id_row, data_start in tab_configs:
        if tab_name not in wb.sheetnames:
            print(f"  Warning: tab '{tab_name}' not found, skipping")
            continue

        ws = wb[tab_name]
        print(f"  Parsing {tab_name}...")

        # Find section columns: cells starting with "P1" in the ID row
        section_cols = []
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=id_row, column=col).value
            if v and str(v).strip().startswith("P1"):
                section_cols.append((col, str(v).strip()))

        # For R3/R5, section IDs are in Row 2
        # Check if they're missing from id_row but present in another row
        if not section_cols:
            for check_row in [1, 2, 3]:
                for col in range(1, ws.max_column + 1):
                    v = ws.cell(row=check_row, column=col).value
                    if v and str(v).strip().startswith("P1"):
                        section_cols.append((col, str(v).strip()))
                if section_cols:
                    break

        print(f"    Found {len(section_cols)} sections: {[s[1] for s in section_cols]}")

        # Find last row with data
        max_data_row = data_start
        for row in range(data_start, min(ws.max_row + 1, 100)):
            has_data = False
            for col in range(2, ws.max_column + 1):
                v = ws.cell(row=row, column=col).value
                if v and str(v).strip():
                    has_data = True
                    break
            if has_data:
                max_data_row = row

        # Parse each section
        for sec_col, sec_id in section_cols:
            entries = []

            # Determine column layout for this section
            # Standard: what=sec_col, qty=sec_col+1, when=sec_col+2, type=sec_col+3
            # Compact (R1.29-39): only 2 columns, what embeds count and type

            # Check if this is a compact section (sec_col+2 has no 'when' header)
            is_compact = False
            header_row = data_start - 1
            when_header = ws.cell(row=header_row, column=sec_col + 2).value
            if when_header is None or (isinstance(when_header, str) and "when" not in when_header.lower()):
                # Check if what column has compact entries
                first_what = ws.cell(row=data_start, column=sec_col).value
                if first_what and re.match(r"^\d+\s+\w+.*-[ps]$", str(first_what).strip(), re.I):
                    is_compact = True

            for row in range(data_start, max_data_row + 1):
                what_raw = ws.cell(row=row, column=sec_col).value
                if not what_raw or not str(what_raw).strip():
                    continue

                what_text = str(what_raw).strip()

                if is_noise(what_text):
                    continue

                if is_compact:
                    # Compact format: "13 tomato-p"
                    parsed = parse_compact_entry(what_text)
                    if parsed:
                        count, farmos_name, entry_type = parsed
                        # Date might be in qty column
                        date_raw = ws.cell(row=row, column=sec_col + 1).value
                        date_str = parse_date_value(date_raw)
                        if farmos_name:
                            entries.append({
                                "species": farmos_name,
                                "count": count,
                                "date": date_str or "2025-10-04",
                                "entry_type": entry_type,
                                "raw_name": what_text,
                            })
                        else:
                            print(f"    ? Unmapped compact: '{what_text}'")
                    continue

                # Standard format: 4 columns
                qty_raw = ws.cell(row=row, column=sec_col + 1).value
                when_raw = ws.cell(row=row, column=sec_col + 2).value
                type_raw = ws.cell(row=row, column=sec_col + 3).value if sec_col + 3 <= ws.max_column else None

                # Normalize species
                farmos_name = normalize_species(what_text)
                if not farmos_name:
                    if not is_noise(what_text):
                        print(f"    ? Unmapped: '{what_text}' in {sec_id}")
                    continue

                # Parse quantity
                count, unit = parse_quantity(qty_raw)

                # Parse date
                date_str = parse_date_value(when_raw)

                # Parse entry type (p=plant, s=seed)
                entry_type = "p"
                if type_raw:
                    type_str = str(type_raw).strip().lower()
                    if type_str.startswith("s"):
                        entry_type = "s"
                # Infer seed type from quantity text ("10 seeds", "50 seeds")
                if entry_type == "p" and unit == "seeds":
                    entry_type = "s"

                entries.append({
                    "species": farmos_name,
                    "count": count,
                    "date": date_str,
                    "entry_type": entry_type,
                    "raw_name": what_text,
                })

            all_sections[sec_id] = entries
            print(f"    {sec_id}: {len(entries)} entries "
                  f"({sum(1 for e in entries if e['entry_type'] == 'p')}p, "
                  f"{sum(1 for e in entries if e['entry_type'] == 's')}s)")

    return all_sections


# ── farmOS importer ──────────────────────────────────────────


class P1Importer:
    """Imports P1 planting data into farmOS."""

    def __init__(self, dry_run=False):
        self.dry_run = dry_run
        self.client = None
        self.plant_type_cache = {}
        self.section_cache = {}
        self.stats = {
            "plants_created": 0,
            "plants_skipped": 0,
            "transplanting_logs": 0,
            "seeding_logs": 0,
            "failed": 0,
        }

    def connect(self):
        load_dotenv()
        config = {
            "hostname": os.getenv("FARMOS_URL"),
            "username": os.getenv("FARMOS_USERNAME"),
            "password": os.getenv("FARMOS_PASSWORD"),
            "client_id": os.getenv("FARMOS_CLIENT_ID", "farm"),
            "scope": os.getenv("FARMOS_SCOPE", "farm_manager"),
        }

        missing = [k for k in ("hostname", "username", "password") if not config[k]]
        if missing:
            print(f"ERROR: Missing env vars: {', '.join('FARMOS_' + k.upper() for k in missing)}")
            sys.exit(1)

        print(f"\nConnecting to {config['hostname']}...")
        self.client = farmOS_client(
            hostname=config["hostname"],
            client_id=config["client_id"],
            scope=config["scope"],
        )
        self.client.authorize(
            username=config["username"],
            password=config["password"],
        )
        print("Connected.")
        return True

    def fetch_by_name(self, api_path, name):
        session = self.client.session
        encoded = urllib.parse.quote(name)
        url = f"/api/{api_path}?filter[name]={encoded}&page[limit]=50"
        resp = session.http_request(url)
        if resp.status_code != 200:
            return []
        return resp.json().get("data", [])

    def get_plant_type_uuid(self, farmos_name):
        if farmos_name in self.plant_type_cache:
            return self.plant_type_cache[farmos_name]
        terms = self.fetch_by_name("taxonomy_term/plant_type", farmos_name)
        if terms:
            uuid = terms[0]["id"]
            self.plant_type_cache[farmos_name] = uuid
            return uuid
        return None

    def get_section_uuid(self, section_id):
        if section_id in self.section_cache:
            return self.section_cache[section_id]
        assets = self.fetch_by_name("asset/land", section_id)
        if assets:
            uuid = assets[0]["id"]
            self.section_cache[section_id] = uuid
            return uuid
        return None

    def plant_asset_exists(self, asset_name):
        assets = self.fetch_by_name("asset/plant", asset_name)
        return assets[0]["id"] if assets else None

    def format_planted_label(self, date_str):
        if not date_str:
            return "SPRING 2025"
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            return dt.strftime("%-d %b %Y").upper()
        except ValueError:
            return date_str.upper()

    def date_to_timestamp(self, date_str):
        if not date_str:
            return int(datetime(2025, 10, 1, 9, 0, tzinfo=AEST).timestamp())
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d").replace(
                hour=9, tzinfo=AEST
            )
            return int(dt.timestamp())
        except ValueError:
            return int(datetime(2025, 10, 1, 9, 0, tzinfo=AEST).timestamp())

    def create_plant_asset(self, name, plant_type_uuid, notes=""):
        data = {
            "attributes": {
                "name": name,
                "status": "active",
            },
            "relationships": {
                "plant_type": {
                    "data": [{"type": "taxonomy_term--plant_type", "id": plant_type_uuid}]
                },
            },
        }
        if notes:
            data["attributes"]["notes"] = {"value": notes, "format": "default"}

        try:
            result = self.client.asset.send("plant", data)
            return result.get("data", {}).get("id")
        except Exception as e:
            print(f"    ! Failed to create plant asset: {e}")
            return None

    def create_quantity(self, plant_id, count):
        payload = {
            "data": {
                "type": "quantity--standard",
                "attributes": {
                    "value": {"decimal": str(count)},
                    "measure": "count",
                    "label": "plants",
                    "inventory_adjustment": "reset",
                },
                "relationships": {
                    "units": {
                        "data": {
                            "type": "taxonomy_term--unit",
                            "id": PLANT_UNIT_UUID,
                        }
                    },
                    "inventory_asset": {
                        "data": {
                            "type": "asset--plant",
                            "id": plant_id,
                        }
                    },
                },
            }
        }
        try:
            session = self.client.session
            resp = session.http_request(
                "/api/quantity/standard",
                method="POST",
                options={"json": payload},
            )
            return resp.json().get("data", {}).get("id")
        except Exception as e:
            print(f"    ! Failed to create quantity: {e}")
            return None

    def create_transplanting_log(self, plant_id, section_uuid, quantity_id,
                                  date_str, section_id, farmos_name, entry_type):
        timestamp = self.date_to_timestamp(date_str)
        action = "Planted" if entry_type == "p" else "Seeded"
        log_name = f"{action} {section_id} — {farmos_name}"

        log_data = {
            "attributes": {
                "name": log_name,
                "timestamp": str(timestamp),
                "status": "done",
                "is_movement": True,
            },
            "relationships": {
                "asset": {
                    "data": [{"type": "asset--plant", "id": plant_id}]
                },
                "location": {
                    "data": [{"type": "asset--land", "id": section_uuid}]
                },
            },
        }
        if quantity_id:
            log_data["relationships"]["quantity"] = {
                "data": [{"type": "quantity--standard", "id": quantity_id}]
            }

        log_type = "transplanting" if entry_type == "p" else "seeding"
        try:
            result = self.client.log.send(log_type, log_data)
            return result.get("data", {}).get("id")
        except Exception as e:
            print(f"    ! Failed to create {log_type} log: {e}")
            return None

    def import_sections(self, parsed_sections, row_filter=None):
        print(f"\n{'='*60}")
        print("FIREFLY CORNER FARM — P1 Importer")
        print(f"{'='*60}")

        if self.dry_run:
            print("\n  DRY RUN — No changes will be made\n")

        # Filter by row
        if row_filter:
            parsed_sections = {
                k: v for k, v in parsed_sections.items()
                if k.startswith(row_filter)
            }
            print(f"  Filtered to {len(parsed_sections)} sections matching {row_filter}")

        # Count totals
        total_entries = sum(len(v) for v in parsed_sections.values())
        all_species = set()
        for entries in parsed_sections.values():
            for e in entries:
                all_species.add(e["species"])

        print(f"  {len(parsed_sections)} sections, {total_entries} entries, "
              f"{len(all_species)} unique species\n")

        if not self.dry_run:
            self.connect()

            # Pre-validate plant types
            print(f"  Validating {len(all_species)} plant types...")
            missing = []
            for sp in sorted(all_species):
                if not self.get_plant_type_uuid(sp):
                    missing.append(sp)
            if missing:
                print(f"\n  ! {len(missing)} plant types not found:")
                for name in missing:
                    print(f"      - {name}")
                print("\n  Add these plant types first, then re-run.")
                return False
            print(f"    All {len(all_species)} verified.")

            # Pre-validate sections
            section_ids = list(parsed_sections.keys())
            print(f"  Validating {len(section_ids)} sections...")
            missing_secs = []
            for sid in sorted(section_ids):
                if not self.get_section_uuid(sid):
                    missing_secs.append(sid)
            if missing_secs:
                print(f"\n  ! {len(missing_secs)} sections not found:")
                for sid in missing_secs:
                    print(f"      - {sid}")
                print("\n  Create these land assets first, then re-run.")
                return False
            print(f"    All {len(section_ids)} verified.\n")

        # Process each section
        for section_id in sorted(parsed_sections.keys()):
            entries = parsed_sections[section_id]
            if not entries:
                continue

            print(f"\n  {section_id}: {len(entries)} entries")
            section_uuid = self.section_cache.get(section_id)

            # Group entries by species (aggregate counts for same species+type)
            species_groups = {}
            for e in entries:
                key = (e["species"], e["entry_type"])
                if key not in species_groups:
                    species_groups[key] = {
                        "species": e["species"],
                        "entry_type": e["entry_type"],
                        "count": 0,
                        "has_count": False,
                        "date": e["date"],
                        "raw_names": [],
                    }
                if e["count"] is not None:
                    species_groups[key]["count"] += e["count"]
                    species_groups[key]["has_count"] = True
                species_groups[key]["raw_names"].append(e["raw_name"])
                # Use earliest date
                if e["date"] and (not species_groups[key]["date"] or e["date"] < species_groups[key]["date"]):
                    species_groups[key]["date"] = e["date"]

            for (species, entry_type), group in sorted(species_groups.items()):
                count = group["count"]
                has_count = group["has_count"]
                date_str = group["date"]
                planted_label = self.format_planted_label(date_str)
                type_marker = "🌱" if entry_type == "p" else "🌰"

                asset_name = f"{planted_label} - {species} - {section_id}"

                # Skip entries with explicit zero count, but keep entries with no count
                # (seeding records that just note "we seeded this" without gram quantities)
                if has_count and count <= 0:
                    continue

                if self.dry_run:
                    count_str = str(count) if has_count else "—"
                    print(f"    {type_marker} {asset_name} (count: {count_str})")
                    self.stats["plants_created"] += 1
                    if entry_type == "p":
                        self.stats["transplanting_logs"] += 1
                    else:
                        self.stats["seeding_logs"] += 1
                    continue

                # Idempotent check
                existing_id = self.plant_asset_exists(asset_name)
                if existing_id:
                    print(f"    = {asset_name} (exists)")
                    self.stats["plants_skipped"] += 1
                    continue

                plant_type_uuid = self.get_plant_type_uuid(species)

                # Build notes
                notes_parts = []
                if len(group["raw_names"]) > 1:
                    notes_parts.append(
                        f"Combined from: {', '.join(group['raw_names'])}"
                    )
                if entry_type == "s":
                    notes_parts.append("Seeded (not transplanted)")
                if not has_count:
                    notes_parts.append("Quantity not recorded")
                notes = "; ".join(notes_parts) if notes_parts else ""

                # 1. Create plant asset
                plant_id = self.create_plant_asset(asset_name, plant_type_uuid, notes)
                if not plant_id:
                    self.stats["failed"] += 1
                    continue
                self.stats["plants_created"] += 1

                # 2. Create quantity (only if we have a count)
                quantity_id = None
                if has_count and count > 0:
                    quantity_id = self.create_quantity(plant_id, count)

                # 3. Create transplanting/seeding log
                log_id = self.create_transplanting_log(
                    plant_id, section_uuid, quantity_id,
                    date_str, section_id, species, entry_type,
                )
                if log_id:
                    print(f"    {type_marker} {asset_name} (count: {count})")
                    if entry_type == "p":
                        self.stats["transplanting_logs"] += 1
                    else:
                        self.stats["seeding_logs"] += 1
                else:
                    print(f"    ~ {asset_name} (plant OK, log failed)")
                    self.stats["failed"] += 1

        # Summary
        print(f"\n{'='*60}")
        print("IMPORT SUMMARY")
        print(f"{'='*60}")
        print(f"  Plants created:       {self.stats['plants_created']}")
        print(f"  Plants skipped:       {self.stats['plants_skipped']}")
        print(f"  Transplanting logs:   {self.stats['transplanting_logs']}")
        print(f"  Seeding logs:         {self.stats['seeding_logs']}")
        print(f"  Failed:               {self.stats['failed']}")

        if self.dry_run:
            print(f"\n  ** DRY RUN — run without --dry-run to apply changes **")
        else:
            print(f"\n  Import completed!")

        return True


def main():
    parser = argparse.ArgumentParser(description="Import P1 field sheet into farmOS")
    parser.add_argument(
        "--xlsx", default=None,
        help=f"Path to P1 spreadsheet (default: {DEFAULT_XLSX} or ~/Downloads/...)",
    )
    parser.add_argument(
        "--row", default=None,
        help="Filter to specific row (e.g., P1R1, P1R3, P1R5, P1ED1)",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Preview import without making changes",
    )
    args = parser.parse_args()

    # Find spreadsheet
    xlsx_path = args.xlsx
    if not xlsx_path:
        if Path(DEFAULT_XLSX).exists():
            xlsx_path = DEFAULT_XLSX
        elif Path(ALT_XLSX).exists():
            xlsx_path = ALT_XLSX
        else:
            print(f"Error: P1 spreadsheet not found at:")
            print(f"  {DEFAULT_XLSX}")
            print(f"  {ALT_XLSX}")
            print(f"\nCopy the file or use --xlsx to specify the path.")
            sys.exit(1)

    print(f"Reading: {xlsx_path}")

    # Parse spreadsheet
    parsed = parse_p1_spreadsheet(xlsx_path)

    # Import
    importer = P1Importer(dry_run=args.dry_run)
    success = importer.import_sections(parsed, row_filter=args.row)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
