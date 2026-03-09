#!/usr/bin/env python3
"""
Parse Claire's field sheet Excel files into structured JSON.

Handles four spreadsheet formats:
- v2 format (P2R2, P2R3): 10-column inventory layout with farmos_name species
- P2R1 format: Old renovation layout with lifecycle columns, species in Col C
- P2R4 format: v2-like but with multiple inventory dates, two count column modes
- P2R5 format: Registration format with Plant/Seed distinction, per-plant dates

Species names are normalized to match the farmos_name convention in plant_types.csv.
Section IDs are extracted from sheet content or tab names depending on format.

Usage:
    python scripts/parse_fieldsheets.py --input fieldsheets/ --output site/src/data/
    python scripts/parse_fieldsheets.py  # uses defaults
"""

import argparse
import csv
import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl

# ─── Section ID remapping (spreadsheet → farmOS) ─────────────────────────
# For sections where the spreadsheet has a different ID than farmOS
SECTION_ID_MAP = {
    "P2R1.13-22": "P2R1.16-25",
}

# ─── Species name mapping ────────────────────────────────────────────────
# Manual overrides for names that can't be auto-resolved
SPECIES_NAME_OVERRIDES = {
    # P2R1 uppercase/informal names
    "CITRUS yuzu": "Citrus (Yuzu)",
    "COMFREY FFC": "Comfrey",
    "CORIANDER FFC": "Coriander",
    "EUCALYPT": "Eucalypt-Gum (Generic)",
    "FIG TREE": "Fig",
    "FINGER LIME": "Finger Lime",
    "GINGER FFC": "Ginger",
    "GRAPE vine": "Grape Vine",
    "GRAPES": "Grape Vine",
    "GUAVA": "Guava",
    "ICE CREAM BEAN": "Ice Cream Bean",
    "LAVENDER": "Lavender",
    "LEMON TREE FFC cut": "Lemon",
    "LEMON TREE FFC seedl": "Lemon",
    "OLIVE TREE": "Olive",
    "OREGANO": "Oregano",
    "PIGEON PEA": "Pigeon Pea",
    "ROSEMARY": "Rosemary",
    "SAGE": "Sage",
    "TANSY": "Tansy",
    "THYME FFC": "Thyme",
    "borage": "Borage",
    "bush mint": "Mint",
    "chives": "Chives",
    "cow pea": "Cowpea",
    "geranium cuttings": "Geranium",
    "hyssop": "Hyssop",
    "lemon balm": "Lemon Balm",
    "lemon scented tea tree cutting": "Tea Tree Oil (Melaleuca) (Alternifolia)",
    "sweet potatoes": "Sweet Potato",
    "Parsley": "Parsley (Italian)",
    "oat": "Oat",
    "millet": "Millet",
    "sunn hemp": "Sunn Hemp",
    # P2R2/P2R3 names that differ from farmos_name
    "Cherry Guava": "Guava (Strawberry)",
    "Cherry Guava (Strawberry)": "Guava (Strawberry)",
    "Cootamundra Wattle": "Wattle - Cootamundra (Baileyana)",
    "Silver Wattle": "Silver Wattle (Dealbata)",
    "White Mulberry": "Mulberry (White)",
    "Basil (Sweet)": "Basil - Sweet (Classic)",
    "Basil (Thai)": "Basil - Perennial (Thai)",
    "Basil (Greek)": "Basil - Perennial (Greek)",
    "Plum": "Plum (Generic)",
    "Pumpkin": "Pumpkin (Generic)",
    "Tea Tree": "Tea Tree Oil (Melaleuca) (Alternifolia)",
    "Tagasaste": "Tagasaste - Tree Lucerne",
    "Tallowood": "Tallowood (Gum)",
    # Green manure species
    "Bean (Climbing) Epicure": "Bean - Climbing",
    "Bean (Dwarf) Borlotti": "Bean - Dwarf",
    "Buck wheat": "Buckwheat",
    "Butternut ffc": "Butternut Pumpkin",
    "Cow pea": "Cowpea",
    "Pea (Snow/Sugar Snap)": "Pea-Snow",
    "Pigeon Pea (seeds)": "Pigeon Pea",
    "Pumpkin ffc": "Pumpkin (Generic)",
    "Sunn hemp": "Sunn Hemp",
    "White Clover": "Clover (White)",
    "Corn": "Corn",
    "Spring Onion": "Spring Onion",
    "Sunflower": "Sunflower",
    # P2R4 species names
    "Basil - Sweet": "Basil - Sweet (Classic)",
    "Tumeric": "Turmeric",
    "Macuya": "Passionfruit",
    "Red Cabbage": "Cabbage (Red)",
    "red cabbage": "Cabbage (Red)",
    "Radish": "Radish (Generic)",
    "radish": "Radish (Generic)",
    "Pepino": "Pepino plant",
    # P2R5 species names (registration format)
    "Ice ream bean": "Ice Cream Bean",
    "Mulberry white": "Mulberry (White)",
    "Thai basil": "Basil - Perennial (Thai)",
    "Cootamundra wattle": "Wattle - Cootamundra (Baileyana)",
    "Parsley Giant Italian": "Parsley (Italian)",
    "Parsley Italian": "Parsley (Italian)",
    "Parsley curled moss": "Parsley (Moss Curled)",
    "Parsley Moss Curled": "Parsley (Moss Curled)",
    "Apple tree seedling": "Apple",
    "Avocado tree seedling": "Avocado",
    "Butternut Pumpkin": "Butternut Pumpkin",
    "Dianella - australian flax": "Dianella",
    "Prickly-leaved paperbark or prickly tea-tree": "Prickly-leaved Paperbark",
    "Prickly-leaved paperbark, prickly tea-tree": "Ball Honey Myrtle",
    "Flax-leaved paperbark, snow in summer": "Snow in Summer (Melaleuca) (Linariifolia)",
    "White Feather, Honey Myrtle": "Melaleuca-White Feather Honey Myrtle (Decora)",
    "Black she-oak": "Black She-oak",
    "Pigeon pea seeds": "Pigeon Pea",
    "Tallowood tree": "Tallowood (Gum)",
    # Green manure species from P2R4 pres / P2R5 autumn overview tabs
    "broad-bean": "Broad Bean",
    "Broad-bean": "Broad Bean",
    "broad bean": "Broad Bean",
    "pea bush Sugar Snap Bon": "Pea-Sugar Snap",
    "snowpea Yakumo": "Pea-Snow (Yakumo)",
    "pumkin jap": "Pumpkin (Generic)",
    "pumpkin jap": "Pumpkin (Generic)",
    "butternut": "Butternut Pumpkin",
    "RED cabbage": "Cabbage (Red)",
    "vetch - inoculated": "Vetch",
    "choko": "Choko",
    "barley": "Barley",
    "coriander": "Coriander",
    "Ice-cream bean": "Ice Cream Bean",
    "Ice ream bean": "Ice Cream Bean",
}

# Suffixes to strip during normalization
STRIP_SUFFIXES = [" FFC", " ffc", " tree", " Tree", " vine", " cuttings", " cutting", " seedl",
                  " tree seedling", " seedling", " plant"]

# Non-plant rows to skip
SKIP_SPECIES = {
    "LIME g/m2", "lime g/m2", "LIME", "lime", "TOTALS", "totals",
    "GM TOTAL", "gm total", "GREEN MANURE", "green manure",
    "None", "", "SUMMER GREEN MANURE", "WINTER GREEN MANURE",
    "1 lm= 1sqm", "What", "what",
}


def load_farmos_names(csv_path="knowledge/plant_types.csv"):
    """Load all farmos_name values from plant_types.csv for matching."""
    names = set()
    csv_file = Path(csv_path)
    if not csv_file.exists():
        print(f"  Warning: {csv_path} not found, name matching disabled")
        return names
    with open(csv_file, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get("farmos_name", "").strip()
            if name:
                names.add(name)
    return names


def normalize_species(raw_name, farmos_names):
    """Normalize a species name to match farmos_name convention.

    Resolution order:
    1. Exact match in overrides dict
    2. Stripped/trimmed exact match in farmos_names
    3. Case-insensitive match in farmos_names
    4. Strip common suffixes and try again
    5. Title-case and try
    6. Return raw name with flag if no match
    """
    if not raw_name or raw_name.strip() in SKIP_SPECIES:
        return None

    name = raw_name.strip()

    # 1. Manual override (exact)
    if name in SPECIES_NAME_OVERRIDES:
        return SPECIES_NAME_OVERRIDES[name]

    # 2. Direct match in farmos_names
    if name in farmos_names:
        return name

    # 3. Case-insensitive match
    name_lower = name.lower()
    for fn in farmos_names:
        if fn.lower() == name_lower:
            return fn

    # 4. Strip suffixes and retry
    stripped = name
    for suffix in STRIP_SUFFIXES:
        if stripped.endswith(suffix):
            stripped = stripped[:-len(suffix)].strip()
    if stripped != name:
        if stripped in farmos_names:
            return stripped
        for fn in farmos_names:
            if fn.lower() == stripped.lower():
                return fn

    # 5. Title case
    titled = name.strip().title()
    if titled in farmos_names:
        return titled

    # 6. No match found — return as-is (will be flagged)
    return name


def parse_count(val):
    """Parse a numeric count from a cell value."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        n = float(val)
        return int(n) if n == int(n) else n
    try:
        n = float(str(val).strip())
        return int(n) if n == int(n) else n
    except (ValueError, TypeError):
        return None


def format_date(val):
    """Format a datetime or string value as YYYY-MM-DD."""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        m = re.search(r'(\d{4}-\d{2}-\d{2})', val)
        if m:
            return m.group(1)
    return ""


# ─── v2 format parser (P2R2, P2R3) ──────────────────────────────────────

def parse_v2_section(ws, farmos_names):
    """Parse a v2 format section sheet (P2R2/P2R3 inventory format).

    Row 1: "P2 — R{n} — Section {start}-{end} (was {old})"
    Row 2: "{length}m | WITH/NO TREES | First planted: {date} | {climate}"
    Row 3: Length: {n}  Date: {first_planted} {last_inventory}
    Row 4: Headers
    Row 5+: Plant data (Col A=strata, B=species, C=notes, D=planted, E=last inventory)
    """
    title = str(ws.cell(1, 1).value or "")
    meta = str(ws.cell(2, 1).value or "")

    # Extract section range from title — use FIRST range (before any "(was ...)")
    title_clean = title.replace("\u2014", "-").replace("\u2013", "-").replace("—", "-").replace("–", "-")
    m = re.search(r'P(\d+)\s*-\s*R(\d+)\s*-\s*Section\s+([\d.]+-[\d.]+)', title_clean)
    if not m:
        return None

    paddock = int(m.group(1))
    row = int(m.group(2))
    section_range = m.group(3)
    section_id = f"P{paddock}R{row}.{section_range}"

    # Remap if needed
    section_id = SECTION_ID_MAP.get(section_id, section_id)

    # Metadata from Row 2
    has_trees = "WITH TREES" in meta.upper()
    first_planted_m = re.search(r'First planted:\s*([^|]+)', meta, re.IGNORECASE)
    first_planted = first_planted_m.group(1).strip() if first_planted_m else ""

    # Structured data from Row 3
    length_val = ws.cell(3, 2).value  # B3
    length = f"{int(length_val)}m" if length_val else ""
    inventory_date = format_date(ws.cell(3, 5).value)  # E3
    first_planted_date = format_date(ws.cell(3, 4).value)  # D3 — exact date
    if first_planted_date:
        first_planted = first_planted_date  # Prefer exact date over text

    # Parse plant data from Row 5+
    plants = []
    green_manure = []
    current_strata = None
    in_green_manure = False
    unmapped = []

    for row_idx in range(5, ws.max_row + 1):
        a_val = str(ws.cell(row_idx, 1).value or "").strip()
        b_val = str(ws.cell(row_idx, 2).value or "").strip()

        # Detect end of plant data / start of green manure
        if a_val.upper().startswith("TOTALS"):
            continue
        if "GREEN MANURE" in a_val.upper():
            in_green_manure = True
            continue
        if a_val.upper().startswith("GM TOTAL"):
            continue
        if "Firefly Corner Farm" in a_val:
            break

        if in_green_manure:
            # Green manure row: A=Placenta, B=species, E=last seeded quantity
            if not b_val or b_val.lower() in ("species", ""):
                continue
            gm_name = normalize_species(b_val, farmos_names)
            if gm_name:
                gm_qty = parse_count(ws.cell(row_idx, 5).value)
                green_manure.append({
                    "species": gm_name,
                    "quantity_grams": gm_qty,
                })
            continue

        # Regular plant data
        if not b_val:
            continue

        # Update strata
        if a_val:
            a_lower = a_val.lower()
            if a_lower in ("emergent", "high", "medium", "low"):
                current_strata = a_lower
            elif a_lower == "placenta":
                continue  # Skip inline placenta markers

        if not current_strata:
            continue

        # Normalize species name
        species = normalize_species(b_val, farmos_names)
        if not species:
            continue

        if species not in farmos_names and species == b_val:
            unmapped.append(b_val)

        notes = str(ws.cell(row_idx, 3).value or "").strip()
        count = parse_count(ws.cell(row_idx, 5).value)  # Col E = Last Inventory

        plants.append({
            "species": species,
            "strata": current_strata,
            "count": count,
            "notes": notes if notes and notes != "None" else "",
        })

    if unmapped:
        for name in unmapped:
            print(f"    ⚠ Unmapped species: '{name}'")

    result = {
        "id": section_id,
        "paddock": paddock,
        "row": row,
        "range": section_range.replace("-", "\u2013"),
        "length": length,
        "has_trees": has_trees,
        "first_planted": first_planted,
        "inventory_date": inventory_date,
        "plants": plants,
    }
    if green_manure:
        result["green_manure"] = green_manure
    return result


# ─── P2R1 format parser ─────────────────────────────────────────────────

def parse_r1_section(ws, farmos_names):
    """Parse a P2R1 renovation format section sheet.

    Tab names: R1.{range}.2025 spring renovation
    Row 2: H2='LENGTH m', I2=length, J2=inventory date
    Row 3: H3='Portion name', I3=range, J3=farmOS section ID
    Row 4: Headers (Lifecycle columns)
    Row 5: Sub-headers (strata, % surface)
    Row 6+: Data (Col A=strata, Col C=species, Col M=total count)
    """
    # Section ID from J3
    section_id = str(ws.cell(3, 10).value or "").strip()  # J3
    if not section_id or not section_id.startswith("P2R1"):
        return None

    # Remap if needed
    section_id = SECTION_ID_MAP.get(section_id, section_id)

    # Extract paddock, row, range from section_id
    m = re.match(r'P(\d+)R(\d+)\.([\d]+-[\d]+)', section_id)
    if not m:
        return None
    paddock = int(m.group(1))
    row = int(m.group(2))
    section_range = m.group(3)

    # Metadata
    length_val = ws.cell(2, 9).value  # I2
    length = f"{int(length_val)}m" if length_val else ""
    inventory_date = format_date(ws.cell(2, 10).value)  # J2
    has_trees = str(ws.cell(4, 9).value or "").strip().upper() == "TREES"  # I4

    # First planted from A1 text — extract and normalize to ISO date
    a1 = str(ws.cell(1, 1).value or "")
    first_planted = ""
    fp_m = re.search(r'PLANTATION/SOWING DATE\s*:\s*(.+)', a1, re.IGNORECASE)
    if fp_m:
        raw_date = fp_m.group(1).strip()
        # Parse "2025-MARCH-20 to 24TH" → "2025-03-20"
        date_m = re.match(r'(\d{4})[/-](\w+)[/-](\d+)', raw_date)
        if date_m:
            month_names = {
                "JANUARY": 1, "FEBRUARY": 2, "MARCH": 3, "APRIL": 4,
                "MAY": 5, "JUNE": 6, "JULY": 7, "AUGUST": 8,
                "SEPTEMBER": 9, "OCTOBER": 10, "NOVEMBER": 11, "DECEMBER": 12,
            }
            month_num = month_names.get(date_m.group(2).upper(), 0)
            if month_num:
                first_planted = f"{date_m.group(1)}-{month_num:02d}-{int(date_m.group(3)):02d}"
            else:
                first_planted = raw_date
        else:
            first_planted = raw_date

    # Parse plant data from Row 6+
    plants = []
    current_strata = None
    unmapped = []

    for row_idx in range(6, ws.max_row + 1):
        a_val = str(ws.cell(row_idx, 1).value or "").strip()
        c_val = str(ws.cell(row_idx, 3).value or "").strip()  # Species in Col C

        if not c_val:
            continue

        # Skip non-plant entries
        if c_val in SKIP_SPECIES or c_val.startswith("SUMMER GREEN") or c_val.startswith("WINTER GREEN"):
            continue
        if c_val.upper().startswith("LIME G/M"):
            continue

        # Update strata from Col A
        if a_val:
            a_lower = a_val.lower().rstrip(".")
            if a_lower in ("emergent", "high", "medium", "low"):
                current_strata = a_lower
            elif a_lower in ("pionneer", "pionneer.", "pioneer", "placenta"):
                continue  # Skip these markers

        if not current_strata:
            continue

        # Normalize species
        species = normalize_species(c_val, farmos_names)
        if not species:
            continue

        if species not in farmos_names and species == c_val:
            unmapped.append(c_val)

        # Count from Col M (total) — fallback to Col J (inventory) or Col I (initial)
        count = parse_count(ws.cell(row_idx, 13).value)  # Col M
        if count is None:
            count = parse_count(ws.cell(row_idx, 10).value)  # Col J
        if count is None:
            count = parse_count(ws.cell(row_idx, 9).value)  # Col I

        plants.append({
            "species": species,
            "strata": current_strata,
            "count": count,
            "notes": "",
        })

    if unmapped:
        for name in unmapped:
            print(f"    ⚠ Unmapped species: '{name}'")

    return {
        "id": section_id,
        "paddock": paddock,
        "row": row,
        "range": section_range.replace("-", "\u2013"),
        "length": length,
        "has_trees": has_trees,
        "first_planted": first_planted,
        "inventory_date": inventory_date,
        "plants": plants,
    }


# ─── P2R4 format parser ────────────────────────────────────────────────

def parse_r4_section(ws, sheet_name, farmos_names):
    """Parse a P2R4 inventory format section sheet.

    Similar to v2 but with key differences:
    - Section ID derived from tab name (A1 can have copy-paste bugs)
    - Multiple inventory dates in row 3 (use latest)
    - Two count column modes: tabs use Col E or Col F for current count
    - GREENMANURE and COMPANION PLANT strata types
    - Col I (New TOTAL) has the most up-to-date count
    """
    # Section ID from tab name (authoritative — A1 can have bugs)
    tab_m = re.match(r'(P2R4\.\d+-\d+)', sheet_name)
    if not tab_m:
        return None
    section_id = tab_m.group(1)

    m = re.match(r'P(\d+)R(\d+)\.([\d]+-[\d]+)', section_id)
    if not m:
        return None
    paddock = int(m.group(1))
    row = int(m.group(2))
    section_range = m.group(3)

    # Metadata from Row 2
    meta = str(ws.cell(2, 1).value or "")
    has_trees = "WITH TREES" in meta.upper()
    first_planted_m = re.search(r'First planted:\s*([^|]+)', meta, re.IGNORECASE)
    first_planted = ""
    if first_planted_m:
        fp_text = first_planted_m.group(1).strip()
        # Parse "October 2025" → "2025-10-01"
        month_names = {
            "january": 1, "february": 2, "march": 3, "april": 4,
            "may": 5, "june": 6, "july": 7, "august": 8,
            "september": 9, "october": 10, "november": 11, "december": 12,
        }
        fp_m = re.match(r'(\w+)\s+(\d{4})', fp_text)
        if fp_m:
            month_num = month_names.get(fp_m.group(1).lower(), 0)
            if month_num:
                first_planted = f"{fp_m.group(2)}-{month_num:02d}-01"
            else:
                first_planted = fp_text
        else:
            first_planted = fp_text

    # Section length from B3
    length_val = ws.cell(3, 2).value
    length = f"{int(length_val)}m" if length_val else ""

    # Find latest inventory date in row 3 (dates span columns D through I+)
    inventory_date = ""
    latest_date = None
    for col in range(4, min(ws.max_column + 1, 15)):
        val = ws.cell(3, col).value
        if isinstance(val, datetime):
            if latest_date is None or val > latest_date:
                latest_date = val
    if latest_date:
        inventory_date = latest_date.strftime("%Y-%m-%d")

    # Parse plant data from Row 5+
    plants = []
    green_manure = []
    current_strata = None
    in_green_manure = False
    in_placenta = False
    unmapped = []

    for row_idx in range(5, ws.max_row + 1):
        a_val = str(ws.cell(row_idx, 1).value or "").strip()
        b_val = str(ws.cell(row_idx, 2).value or "").strip()

        # Detect end markers
        if a_val.upper().startswith("TOTALS") or "Firefly Corner Farm" in a_val:
            in_green_manure = False
            in_placenta = True  # After TOTALS, Placenta section follows
            continue
        if in_placenta:
            # Skip Placenta (historical green manure) section
            if a_val.upper().startswith("PLACENTA") or b_val:
                continue
            continue

        if not b_val:
            continue

        # Update strata / detect green manure section
        if a_val:
            a_lower = a_val.lower().strip()
            if a_lower in ("emergent", "high", "medium", "low"):
                current_strata = a_lower
                in_green_manure = False
            elif a_lower in ("greenmanure", "green manure"):
                in_green_manure = True
                continue
            elif a_lower in ("companion plant", "companion"):
                in_green_manure = True  # Treat companion plants as green manure
                continue

        if in_green_manure:
            gm_name = normalize_species(b_val, farmos_names)
            if gm_name:
                gm_qty = parse_count(ws.cell(row_idx, 8).value)  # Col H = seeds (g)
                green_manure.append({
                    "species": gm_name,
                    "quantity_grams": gm_qty,
                })
            continue

        if not current_strata:
            continue

        # Normalize species name
        species = normalize_species(b_val, farmos_names)
        if not species:
            continue

        if species not in farmos_names and species == b_val:
            unmapped.append(b_val)

        # Count: try Col I (New TOTAL) first, then Col F (new inventory), then Col E
        count = parse_count(ws.cell(row_idx, 9).value)   # Col I = New TOTAL
        if count is None:
            count = parse_count(ws.cell(row_idx, 6).value)  # Col F = New inventory
        if count is None:
            count = parse_count(ws.cell(row_idx, 5).value)  # Col E = Last inventory

        notes = str(ws.cell(row_idx, 10).value or "").strip()  # Col J = Comments

        plants.append({
            "species": species,
            "strata": current_strata,
            "count": count,
            "notes": notes if notes and notes != "None" else "",
        })

    if unmapped:
        for name in unmapped:
            print(f"    ⚠ Unmapped species: '{name}'")

    result = {
        "id": section_id,
        "paddock": paddock,
        "row": row,
        "range": section_range.replace("-", "\u2013"),
        "length": length,
        "has_trees": has_trees,
        "first_planted": first_planted,
        "inventory_date": inventory_date,
        "plants": plants,
    }
    if green_manure:
        result["green_manure"] = green_manure
    return result


# ─── P2R5 format parser (registration format) ─────────────────────────

def parse_r5_section(ws, sheet_name, farmos_names):
    """Parse a P2R5 registration format section sheet.

    Tab names: P2R5.{range}.JANV.2026
    Row 1: Headers with inventory date in Col I
    Row 2: Section range in Col B, length in Col C, trees in Col D
    Row 3: Column headers
    Row 4+: Data (Col A=species, Col C=strata, Col E=P/S, Col I=count)
    """
    # Section ID from tab name
    tab_m = re.match(r'(P2R5)\.([\d]+-[\d]+)', sheet_name)
    if not tab_m:
        return None

    section_range = tab_m.group(2)
    section_id = f"P2R5.{section_range}"

    # Remap if needed
    section_id = SECTION_ID_MAP.get(section_id, section_id)

    m = re.match(r'P(\d+)R(\d+)\.([\d]+-[\d]+)', section_id)
    if not m:
        return None
    paddock = int(m.group(1))
    row = int(m.group(2))
    section_range = m.group(3)

    # Section length from C2
    length_val = ws.cell(2, 3).value
    length = f"{int(length_val)}m" if length_val else ""

    # Has trees from D2 (all R5 sections say "with")
    trees_val = str(ws.cell(2, 4).value or "").strip().lower()
    has_trees = trees_val in ("with", "trees", "with trees")

    # Inventory date from I1
    inv_date_val = ws.cell(1, 9).value
    inventory_date = format_date(inv_date_val) if inv_date_val else ""

    # Parse plant data from Row 4+
    plants = []
    unmapped = []
    earliest_planted = None

    for row_idx in range(4, ws.max_row + 1):
        a_val = str(ws.cell(row_idx, 1).value or "").strip()  # Species name
        c_val = str(ws.cell(row_idx, 3).value or "").strip()  # Strata
        e_val = str(ws.cell(row_idx, 5).value or "").strip()  # P or S
        i_val = ws.cell(row_idx, 9).value  # Inventory count

        if not a_val or a_val in SKIP_SPECIES:
            continue

        # Skip seed entries — only include plants (P) for landing pages
        if e_val.upper() == "S":
            continue

        # Skip non-plant rows (check for header-like content)
        if a_val.lower() in ("what", "species"):
            continue

        # Normalize strata
        strata = None
        if c_val:
            c_lower = c_val.lower().strip()
            if c_lower == "emergent":
                strata = "emergent"
            elif c_lower == "high":
                strata = "high"
            elif c_lower in ("medium-high", "medium high"):
                strata = "high"  # Map medium-high → high
            elif c_lower == "medium":
                strata = "medium"
            elif c_lower == "low":
                strata = "low"

        # Normalize species name
        species = normalize_species(a_val, farmos_names)
        if not species:
            continue

        if species not in farmos_names and species == a_val:
            unmapped.append(a_val)

        count = parse_count(i_val)  # Col I = inventory count

        # Track earliest planting date for first_planted
        h_val = ws.cell(row_idx, 8).value  # Col H = planting date
        if isinstance(h_val, datetime):
            if earliest_planted is None or h_val < earliest_planted:
                earliest_planted = h_val

        plants.append({
            "species": species,
            "strata": strata,
            "count": count,
            "notes": "",
        })

    if unmapped:
        for name in unmapped:
            print(f"    ⚠ Unmapped species: '{name}'")

    first_planted = ""
    if earliest_planted:
        first_planted = earliest_planted.strftime("%Y-%m-%d")

    return {
        "id": section_id,
        "paddock": paddock,
        "row": row,
        "range": section_range.replace("-", "\u2013"),
        "length": length,
        "has_trees": has_trees,
        "first_planted": first_planted,
        "inventory_date": inventory_date,
        "plants": plants,
    }


# ─── Gap section parsers (overview/autumn tabs) ────────────────────────

def parse_r5_autumn_sections(ws, farmos_names, existing_section_ids):
    """Parse the P2R5 autumn seeds&plants tab for gap (no-tree) sections.

    This tab has a matrix format:
    - Row 4: Section ranges in columns C-J (0-7, 7-14, 14-21, 21-29, 29-37, 37-44, 44-53, 53-80)
    - Row 3: Tree info per column (trees, under 2 m high)
    - Row 5: Label row (Green manure may-june 2025)
    - Rows 6+: Species name in Col A, quantities per section in columns C-J

    Only creates sections for gaps NOT already in existing_section_ids.
    Gap sections are identified by "under 2 m high" in Row 3.
    """
    # Map column index to section info from Row 4 (ranges) and Row 3 (tree info)
    col_sections = {}
    for col in range(3, 11):  # C=3 through J=10
        range_val = str(ws.cell(4, col).value or "").strip()
        tree_val = str(ws.cell(3, col).value or "").strip().lower()
        if not range_val:
            continue
        # Normalize range: " 7-14" → "7-14"
        range_val = range_val.strip()
        is_no_tree = "under" in tree_val or "2 m" in tree_val
        col_sections[col] = {
            "range": range_val,
            "has_trees": not is_no_tree,
        }

    sections = []
    for col, sec_info in col_sections.items():
        # Build the section ID using OUR boundary convention
        # The autumn tab uses original boundaries (0-7, 7-14, etc.)
        # Our sections use adjusted boundaries based on existing sections
        orig_range = sec_info["range"]
        parts = orig_range.split("-")
        if len(parts) != 2:
            continue
        orig_start, orig_end = int(parts[0]), int(parts[1])

        # Only process no-tree gap sections
        if sec_info["has_trees"]:
            continue

        # Find our section boundary by looking at adjacent existing sections
        # E.g., if 0-8 exists, gap section 7-14 becomes 8-14
        our_start = orig_start
        our_end = orig_end
        for sid in existing_section_ids:
            if not sid.startswith("P2R5."):
                continue
            sid_parts = sid.split(".")[1].split("-")
            sid_end = int(sid_parts[1])
            sid_start = int(sid_parts[0])
            # Adjust start if an existing section's end falls within our range
            if sid_end > our_start and sid_end <= our_end:
                our_start = sid_end
            # Adjust end if an existing section's start falls within our range
            if sid_start >= our_start and sid_start < our_end:
                our_end = sid_start

        section_id = f"P2R5.{our_start}-{our_end}"
        if section_id in existing_section_ids:
            continue  # Already parsed from registration tabs

        length = our_end - our_start
        if length < 4:
            continue  # Skip tiny gaps or non-cultivable row ends

        # Parse green manure species from this column
        green_manure = []
        plants = []
        for row_idx in range(6, ws.max_row + 1):
            a_val = str(ws.cell(row_idx, 1).value or "").strip()
            qty_val = ws.cell(row_idx, col).value

            if not a_val:
                continue

            species = normalize_species(a_val, farmos_names)
            if not species:
                continue

            qty = parse_count(qty_val)
            if qty is None or qty == 0:
                continue

            # Choko is a plant (vine), not a green manure seed
            if species == "Choko":
                plants.append({
                    "species": species,
                    "strata": "medium",
                    "count": int(qty),
                    "notes": "Sowed May-June 2025",
                })
            else:
                green_manure.append({
                    "species": species,
                    "quantity_grams": qty,
                })

        if not green_manure and not plants:
            continue

        section = {
            "id": section_id,
            "paddock": 2,
            "row": 5,
            "range": f"{our_start}\u2013{our_end}",
            "length": f"{length}m",
            "has_trees": False,
            "first_planted": "2025-05-28",  # "sowed + lime from MAY 28TH to JUNE 5TH"
            "inventory_date": "",  # No inventory done for these sections
            "plants": plants,
        }
        if green_manure:
            section["green_manure"] = green_manure
        sections.append(section)
        print(f"  Parsed gap: {section_id} — {len(plants)} plants, {len(green_manure)} green manure species")

    return sections


def parse_r4_pres_sections(ws, farmos_names, existing_section_ids):
    """Parse the P2R4 pres/BEFORE tab for gap (no-tree) sections.

    This tab has a matrix format:
    - Row 3: Section ranges in columns C-M (0-1, 1-6, 6-13, 13-20, ...)
    - Row 2: Section lengths
    - Rows 5-14: Green manure species (merged cell C5:H14 = "GO TO FARM-OS")
    - Col B: Seed density per m² (used to calculate quantities for gap sections)

    Gap sections L6 (1-6) and L8 (13-20) don't have explicit quantities.
    We calculate them from seed density × section length.
    """
    # Map column index to section info from Row 3 (ranges)
    col_sections = {}
    for col in range(3, 14):  # C=3 through M=13
        range_val = str(ws.cell(3, col).value or "").strip()
        if not range_val or range_val == "PATH":
            continue
        range_val = range_val.strip()
        parts = range_val.split("-")
        if len(parts) != 2:
            continue
        try:
            start, end = int(parts[0]), int(parts[1])
        except ValueError:
            continue
        length_val = ws.cell(2, col).value
        length = int(float(length_val)) if length_val else (end - start)
        col_sections[col] = {
            "range": range_val,
            "start": start,
            "end": end,
            "length": length,
        }

    # Identify gap sections: columns with no explicit green manure quantities
    # These are in the merged cell region (C5:H14 → cols 3-8)
    # We need to identify which of these correspond to "no tree" gap sections
    # P2R4 original: L5(0-1) L6(1-6) L7(6-13) L8(13-20) L9(20-27) ...
    # L6 and L8 are the open sections between tree sections

    # Parse seed densities from Col B
    seed_densities = {}
    for row_idx in range(5, 15):
        a_val = str(ws.cell(row_idx, 1).value or "").strip()
        b_val = ws.cell(row_idx, 2).value
        if not a_val:
            continue

        species = normalize_species(a_val, farmos_names)
        if not species:
            continue

        # Parse density: "60s or 3g" → use seed count, or "20.0" → seeds/m²
        density = None
        if isinstance(b_val, (int, float)):
            density = float(b_val)
        elif isinstance(b_val, str):
            # "60s or 3g" → 3 (use grams)
            gm = re.search(r'(\d+)\s*g', str(b_val))
            if gm:
                density = float(gm.group(1))
            else:
                sm = re.search(r'(\d+)', str(b_val))
                if sm:
                    density = float(sm.group(1))

        if density:
            seed_densities[species] = density

    sections = []

    # Define R4 gap sections (no-tree open cultivation sections)
    # These are between existing sections
    r4_gaps = [
        {"orig_range": "1-6", "orig_start": 1, "orig_end": 6},
        {"orig_range": "13-20", "orig_start": 13, "orig_end": 20},
    ]

    for gap in r4_gaps:
        # Adjust boundaries based on existing sections
        our_start = gap["orig_start"]
        our_end = gap["orig_end"]
        for sid in existing_section_ids:
            if not sid.startswith("P2R4."):
                continue
            sid_parts = sid.split(".")[1].split("-")
            sid_end = int(sid_parts[1])
            sid_start = int(sid_parts[0])
            if sid_end > our_start and sid_end <= our_end:
                our_start = sid_end
            if sid_start >= our_start and sid_start < our_end:
                our_end = sid_start

        section_id = f"P2R4.{our_start}-{our_end}"
        if section_id in existing_section_ids:
            continue

        length = our_end - our_start

        # Calculate green manure quantities from seed density × length
        green_manure = []
        for species, density in seed_densities.items():
            qty = round(density * length, 1)
            if qty > 0:
                green_manure.append({
                    "species": species,
                    "quantity_grams": qty,
                })

        if not green_manure:
            continue

        section = {
            "id": section_id,
            "paddock": 2,
            "row": 4,
            "range": f"{our_start}\u2013{our_end}",
            "length": f"{length}m",
            "has_trees": False,
            "first_planted": "2025-06-01",  # Green manure sowed June 2024, re-sowed summer 2025
            "inventory_date": "",
            "plants": [],
        }
        if green_manure:
            section["green_manure"] = green_manure
        sections.append(section)
        print(f"  Parsed gap: {section_id} — {len(green_manure)} green manure species")

    return sections


# ─── File-level parsing ─────────────────────────────────────────────────

def is_r1_section_tab(name):
    """Check if tab name matches P2R1 section pattern."""
    return bool(re.match(r'R1\.\d+-\d+\.', name))


def is_v2_section_tab(name):
    """Check if tab name matches v2 section pattern (P2R2.x-y or P2R3.x-y)."""
    return bool(re.match(r'P2R[23]\.\d+-\d+', name))


def is_r4_section_tab(name):
    """Check if tab name matches P2R4 inventory section pattern (exactly P2R4.X-Y)."""
    return bool(re.fullmatch(r'P2R4\.\d+-\d+', name))


def is_r5_section_tab(name):
    """Check if tab name matches P2R5 section pattern."""
    return bool(re.match(r'P2R5\.\d+-\d+', name))


def is_r5_autumn_tab(name):
    """Check if tab name matches P2R5 autumn seeds & plants tab."""
    return "automn" in name.lower() or "autumn" in name.lower()


def is_r4_pres_tab(name):
    """Check if tab name is the P2R4 pres/BEFORE overview tab."""
    return name in ("P2R4pres", "P2R4.BEFORE")


def parse_fieldsheet_file(filepath, farmos_names, all_existing_ids=None):
    """Parse all section sheets from a single field sheet Excel file.

    Args:
        all_existing_ids: Set of section IDs already parsed from other files.
                          Used by gap parsers to calculate correct boundaries.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if is_r1_section_tab(sheet_name):
            section = parse_r1_section(ws, farmos_names)
            if section:
                sections.append(section)
                plant_count = sum(1 for p in section["plants"] if p.get("count") and p["count"] > 0)
                print(f"  Parsed: {section['id']} — {len(section['plants'])} species ({plant_count} with count)")
            else:
                print(f"  Skipped: {sheet_name} (could not parse P2R1 format)")

        elif is_v2_section_tab(sheet_name):
            section = parse_v2_section(ws, farmos_names)
            if section:
                sections.append(section)
                plant_count = sum(1 for p in section["plants"] if p.get("count") and p["count"] > 0)
                print(f"  Parsed: {section['id']} — {len(section['plants'])} species ({plant_count} with count)")
            else:
                print(f"  Skipped: {sheet_name} (could not parse v2 format)")

        elif is_r4_section_tab(sheet_name):
            section = parse_r4_section(ws, sheet_name, farmos_names)
            if section:
                sections.append(section)
                plant_count = sum(1 for p in section["plants"] if p.get("count") and p["count"] > 0)
                print(f"  Parsed: {section['id']} — {len(section['plants'])} species ({plant_count} with count)")
            else:
                print(f"  Skipped: {sheet_name} (could not parse P2R4 format)")

        elif is_r5_section_tab(sheet_name):
            section = parse_r5_section(ws, sheet_name, farmos_names)
            if section:
                sections.append(section)
                plant_count = sum(1 for p in section["plants"] if p.get("count") and p["count"] > 0)
                print(f"  Parsed: {section['id']} — {len(section['plants'])} species ({plant_count} with count)")
            else:
                print(f"  Skipped: {sheet_name} (could not parse P2R5 format)")

        else:
            # Skip non-section tabs (mapping, recap, planning, etc.)
            pass

    # Second pass: parse overview/autumn tabs for gap sections
    # Combine IDs from this file AND all previously parsed files
    existing_ids = {s["id"] for s in sections}
    if all_existing_ids:
        existing_ids |= all_existing_ids

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if is_r5_autumn_tab(sheet_name):
            gap_sections = parse_r5_autumn_sections(ws, farmos_names, existing_ids)
            for sec in gap_sections:
                sections.append(sec)
                existing_ids.add(sec["id"])

        elif is_r4_pres_tab(sheet_name):
            gap_sections = parse_r4_pres_sections(ws, farmos_names, existing_ids)
            for sec in gap_sections:
                sections.append(sec)
                existing_ids.add(sec["id"])

    return sections


def build_row_info(sections):
    """Build row-level summary from parsed sections."""
    rows = {}
    for sec in sections:
        row_id = f"P{sec['paddock']}R{sec['row']}"
        if row_id not in rows:
            rows[row_id] = {
                "id": row_id,
                "paddock": f"Paddock {sec['paddock']}",
                "row": f"Row {sec['row']}",
                "sections": [],
                "first_planted": sec.get("first_planted", ""),
            }
        rows[row_id]["sections"].append(sec["id"])

    # Sort sections within each row by start position
    for row in rows.values():
        row["sections"].sort(key=lambda sid: float(sid.split(".")[-1].split("\u2013")[0].split("-")[0]))
        # Calculate total length from last section endpoint
        last_sec = [s for s in sections if s["id"] == row["sections"][-1]][0]
        range_parts = last_sec["range"].replace("\u2013", "-").split("-")
        if len(range_parts) == 2:
            row["total_length"] = f"{range_parts[1]}m"

    return rows


def main():
    parser = argparse.ArgumentParser(description="Parse field sheet Excel files into JSON")
    parser.add_argument("--input", default="fieldsheets/", help="Directory containing .xlsx files")
    parser.add_argument("--output", default="site/src/data/", help="Output directory for JSON")
    parser.add_argument("--plants-csv", default="knowledge/plant_types.csv",
                        help="Plant types CSV for name matching")
    args = parser.parse_args()

    input_dir = Path(args.input)
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Load farmos_names for species matching
    farmos_names = load_farmos_names(args.plants_csv)
    print(f"Loaded {len(farmos_names)} plant type names for matching")

    all_sections = []
    all_existing_ids = set()

    # Find and parse all field sheet files
    xlsx_files = sorted(input_dir.glob("*.xlsx"))
    if not xlsx_files:
        print(f"No .xlsx files found in {input_dir}")
        return

    for filepath in xlsx_files:
        print(f"\nParsing: {filepath.name}")
        sections = parse_fieldsheet_file(filepath, farmos_names, all_existing_ids)
        all_sections.extend(sections)
        all_existing_ids |= {s["id"] for s in sections}

    if not all_sections:
        print("\nNo sections parsed!")
        return

    # Build row summaries
    rows = build_row_info(all_sections)

    # Collect all species for summary
    all_species = set()
    for s in all_sections:
        for p in s["plants"]:
            all_species.add(p["species"])

    unmatched = all_species - farmos_names
    matched = all_species & farmos_names

    # Write sections.json
    output_file = output_dir / "sections.json"
    output_data = {
        "generated": True,
        "sections": {s["id"]: s for s in all_sections},
        "rows": rows,
    }

    with open(output_file, "w") as f:
        json.dump(output_data, f, indent=2, ensure_ascii=False)

    print(f"\n{'='*60}")
    print(f"PARSE SUMMARY")
    print(f"{'='*60}")
    print(f"  Sections parsed:  {len(all_sections)}")
    for row_id, row_info in sorted(rows.items()):
        print(f"    {row_id}: {len(row_info['sections'])} sections")
    print(f"  Unique species:   {len(all_species)}")
    print(f"  Matched to CSV:   {len(matched)}")
    if unmatched:
        print(f"  UNMATCHED:        {len(unmatched)}")
        for name in sorted(unmatched):
            print(f"    - {name}")
    print(f"\n  Output: {output_file}")


if __name__ == "__main__":
    main()
