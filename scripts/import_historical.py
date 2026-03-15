#!/usr/bin/env python3
"""
Import historical planting and observation data from Spring 2025 renovation
spreadsheets into farmOS as backdated logs on EXISTING plant assets.

Phase H1: Creates three types of logs per plant:
1. Transplanting log — initial planting (March–May 2025)
2. Observation log — mid-season inventory (Oct/Nov 2025)
3. Transplanting log — renovation additions (Oct/Nov 2025)

No new plant assets are created. The script matches historical species data
to existing plant assets by reconstructing their name from sections.json.

Usage:
    python scripts/import_historical.py --dry-run
    python scripts/import_historical.py --row P2R1 --dry-run
    python scripts/import_historical.py --section P2R2.0-3 --dry-run
    python scripts/import_historical.py --row P2R1
"""

import argparse
import json
import os
import re
import sys
import urllib.parse
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

try:
    from farmOS import farmOS as farmOS_client
except ImportError:
    print("ERROR: farmOS library not installed!")
    sys.exit(1)

# Import species normalization from parse_fieldsheets
from parse_fieldsheets import (
    SPECIES_NAME_OVERRIDES, SKIP_SPECIES, STRIP_SUFFIXES,
    normalize_species, load_farmos_names,
)

PLANT_UNIT_UUID = "2371b79e-a87b-4152-b6e4-ea6a9ed37fd0"
AEST = timezone(timedelta(hours=10))

# ── Section boundary mapping: historical → current ──────────────────────

HISTORICAL_TO_CURRENT = {
    # R1
    "P2R1.0-3":   "P2R1.0-3",
    "P2R1.3-9":   "P2R1.3-9",
    "P2R1.9-13":  "P2R1.9-13",
    "P2R1.13-22": "P2R1.16-25",
    # R2
    "P2R2.0-3":   "P2R2.0-3",
    "P2R2.3-7":   "P2R2.3-9",
    "P2R2.7-16":  "P2R2.9-16",
    "P2R2.16-23": "P2R2.16-23",
    "P2R2.23-26": "P2R2.23-26",
    "P2R2.28-37": "P2R2.28-38",
    "P2R2.37-46": "P2R2.38-46",
    # R3
    "P2R3.0-3":   "P2R3.0-2",
    "P2R3.3-9":   "P2R3.2-9",
    "P2R3.9-14":  "P2R3.9-15",
    "P2R3.15-21": "P2R3.15-21",
    "P2R3.21-26": None,  # no current section
    "P2R3.26-37": "P2R3.26-38",
    "P2R3.41-63": ["P2R3.40-50", "P2R3.50-62"],
    # R4 (spring planting tabs use different boundaries than current)
    "P2R4.0-1":   "P2R4.0-2",
    "P2R4.6-13":  "P2R4.6-14",
    "P2R4.20-27": "P2R4.20-30",
    "P2R4.35-44": "P2R4.40-49",
    # R5 (Kolala Day + early plantings map to current sections)
    "P2R5.0-8":   "P2R5.0-8",
    "P2R5.14-22": "P2R5.14-22",
    "P2R5.29-38": "P2R5.29-38",
}

# ── Additional species overrides for renovation spreadsheets ────────────

HISTORICAL_SPECIES_OVERRIDES = {
    # R2 renovation names
    "CHILLI JALAPENO": "Chilli (Jalapeño)",
    "CHILLI BIRDEYE": "Chilli (Bird's Eye)",
    "CHILLI Jalapeno": "Chilli (Jalapeño)",
    "PLUM TREE": "Plum (Generic)",
    "FFC CITRUS seedling": "Lemon",
    "FFC CITRUS cutting": "Lemon",
    "TOMATO": "Tomato (Marmande)",
    "TUMERIC": "Turmeric",
    "SWEET POTATOES": "Sweet Potato",
    "SWEET POTATOES / potatoes": "Sweet Potato",
    "CAPE GOOSEBERRY": "Cape Gooseberry",
    "MACADAMIA": "Macadamia",
    "ROSE APPLE": "Rose Apple",
    "APPLE": "Apple",
    "PEAR": "Pear (Williams)",
    "GERANIUM CUTTING": "Geranium",
    "GERANIUM Cutting": "Geranium",
    "CALENDULA": "Calendula",
    "pumpkin": "Pumpkin (Generic)",
    "PUMPKIN": "Pumpkin (Generic)",
    "STRAWBERRY": "Strawberry",
    "BLUEBERRY": "Blueberry",
    "RASPBERRY": "Raspberry",
    "Blackberry": "Blackberry",
    "BLACKBERRY": "Blackberry",
    # Tree lucerne
    "tree lucerne seedling ffc": "Tagasaste - Tree Lucerne",
    "tree lucerne": "Tagasaste - Tree Lucerne",
    "Tree Lucerne": "Tagasaste - Tree Lucerne",
    "TREE LUCERNE": "Tagasaste - Tree Lucerne",
    # Carob
    "Carob tree": "Carob",
    "CAROB": "Carob",
    # Apple seedling
    "APPLE seedling ffc": "Apple",
    "Apple seedling ffc": "Apple",
    "APPLE ffc seedling": "Apple",
    # Mulberry
    "Mulberry White": "Mulberry (White)",
    "MULBERRY white": "Mulberry (White)",
    "MULBERRY-WHITE BECK&DAN": "Mulberry (White)",
    "Mulberry White Beck&Dan": "Mulberry (White)",
    "MULBERRY White (Beck&Dan)": "Mulberry (White)",
    "MULBERRY ffc seedling": "Mulberry (White)",
    "MULBERRY (3=df)": "Mulberry (Beenleigh Black)",
    # Capsicum / chilli
    "red capsicum": "Capsicum (Red)",
    "RED CAPSICUM": "Capsicum (Red)",
    "CHILLI Jalapeno (+birdeye +thai +...)": "Chilli (Jalapeño)",
    "CHILLI/ CAPSICUM or Indian Curry": "Chilli (Jalapeño)",
    # Eggplant
    "EggPlant": "Eggplant",
    "eggplant": "Eggplant",
    "EGGPLANT": "Eggplant",
    # Basil
    "basil thai": "Basil - Perennial (Thai)",
    "BASIL THAI": "Basil - Perennial (Thai)",
    "basil thai perennial": "Basil - Perennial (Thai)",
    "PERENNIAL GREEK BASIL": "Basil - Perennial (Greek)",
    "basil sweet": "Basil - Sweet (Classic)",
    "Basil sweet": "Basil - Sweet (Classic)",
    "BASIL SWEET": "Basil - Sweet (Classic)",
    # Pak Choi
    "cabbage Pak choi": "Pak Choi",
    "CABBAGE PAK CHOI": "Pak Choi",
    "cabbage pak choi & red": "Pak Choi",
    # R3 specific
    "GUAVA STRAWBERRY": "Guava (Strawberry)",
    "Guava Strawberry": "Guava (Strawberry)",
    "QUINCE CUTTING FFC": "Quince",
    "QUINCE FFC CUTTING": "Quince",
    "QUINCE CUTTINGS": "Quince",
    "Quince Cutting FFC": "Quince",
    "PIGEON PEA or Ice Cream bean": "Pigeon Pea",
    "PIGEON PEA or?": "Pigeon Pea",
    "CITRUS SEEDLING FFC": "Lemon",
    "CITRUS CUTTING FFC": "Lemon",
    "CITRUS CUTTINGS FFC": "Lemon",
    "Citrus Seedling FFC": "Lemon",
    "STONE FRUIT PEACH": "Peach",
    "STONE FRUIT APRICOT": "Apricot",
    "STONE FRUIT (PEACH, APRICOT or JABOTICABA)": "Jaboticaba",
    "JABOTICABA: stone fruit tree": "Jaboticaba",
    "MALABAR CHESTNUT": "Malabar Chestnut",
    "MALABAR CHESTNUT/ AMLA": "Malabar Chestnut",
    "BANA-GRASS": "Bana Grass",
    "BANAGRASS": "Bana Grass",
    "AMLA": "Amla",
    "LONGAN": "Longan",
    "ACEROLA": "Acerola",
    "TEA TREE": "Tea Tree Oil (Melaleuca) (Alternifolia)",
    "TEA TREE oil": "Tea Tree Oil (Melaleuca) (Alternifolia)",
    "chinese amaranthus": "Amaranth",
    "Chinese Amaranthus": "Amaranth",
    "indian Curry": "Curry Leaf",
    "INDIAN CURRY": "Curry Leaf",
    "GINGER galangal": "Galangal",
    "Ginger Galangal": "Galangal",
    "zucchini": "Zucchini (Blackjack)",
    "ZUCCHINI": "Zucchini (Blackjack)",
    "cabbage": "Cabbage (Golden Acre)",
    "CABBAGE": "Cabbage (Golden Acre)",
    "cabbage red": "Cabbage (Red)",
    "CABBAGE RED": "Cabbage (Red)",
    "tomato": "Tomato (Marmande)",
    "sage pineapple": "Sage",
    "Sage Pineapple": "Sage",
    "nettle": "Nettle",
    "NETTLE": "Nettle",
    "BORAGE": "Borage",
    "Cootamundra Wattle seedling ffc nursery": "Wattle - Cootamundra (Baileyana)",
    "COOTAMUNDRA WATTLE": "Wattle - Cootamundra (Baileyana)",
    "Perennial chives garlic": "Garlic Chives",
    "ARROW (ACHILLEA MILLEFOLIUM)": "Yarrow (White)",
    "JACARANDA": "Jacaranda",
    # Eucalypt variants
    "EUCALYPT-RED FOREST GUM from ffc nursery": "Forest Red Gum",
    "EUCALYPT red forest gum": "Forest Red Gum",
    "EUCALYPT seedling ffc nursery": "Forest Red Gum",
    "Eucalypt seedling ffc nursery": "Forest Red Gum",
    # Ginger with provenance
    "ginger ? Minimbah permaculture farm 22.09.2025": "Ginger",
    "GINGER": "Ginger",
    # Blueberry
    "BLUEBERRY (CUTTINGS AND PLANTS)": "Blueberry",
    # Lavender
    "LAVENDER plants and cuttings": "Lavender",
    # Additional overrides found during dry-run
    "cabbage pak choi": "Pak Choi",
    "cabbage Pak Choi": "Pak Choi",
    "CABBAGE pak choi": "Pak Choi",
    "cabbage pak choi & red": "Pak Choi",
    "YARROW": "Yarrow (White)",
    "Yarrow": "Yarrow (White)",
    "QUINCE CUTTING": "Quince",
    "Quince cutting": "Quince",
    "quince cutting": "Quince",
    "ffc quince cutting": "Quince",
    "FFC QUINCE CUTTING": "Quince",
    "CHILLI-CAPSCICUM": "Chilli (Jalapeño)",
    "Chilli-Capsicum": "Chilli (Jalapeño)",
    "CHILLI/ CAPSICUM": "Chilli (Jalapeño)",
    "FFC CITRUS CUTTING & SEEDLING": "Lemon",
    "ffc citrus cutting & seedling": "Lemon",
    "FFC CITRUS CUTTING & seedling": "Lemon",
    "Citrus seedling nursery ffc": "Lemon",
    "CITRUS seedling nursery ffc": "Lemon",
    "citrus seedling nursery ffc": "Lemon",
    "sweet basil": "Basil - Sweet (Classic)",
    "SWEET BASIL": "Basil - Sweet (Classic)",
    "Sweet Basil": "Basil - Sweet (Classic)",
    "capsicum red": "Capsicum (Red)",
    "CAPSICUM RED": "Capsicum (Red)",
    "Capsicum Red": "Capsicum (Red)",
    "pumpkin ffc": "Pumpkin (Generic)",
    "PUMPKIN FFC": "Pumpkin (Generic)",
    "Pumpkin FFC": "Pumpkin (Generic)",
    "pumpkin FFC": "Pumpkin (Generic)",
    # Multi-species entries — map to primary species
    "1 ORANGE 1 GRAPEFRUIT 1 KUMQUAT": None,  # skip: can't map to single species
    # Generic Guava → Guava (Strawberry) (only variety on farm)
    "GUAVA": "Guava (Strawberry)",
    "Guava": "Guava (Strawberry)",
    "guava": "Guava (Strawberry)",
    # R4 spring planting tab names
    "Sweet potato": "Sweet Potato",
    "tree lucerne": "Tagasaste - Tree Lucerne",
    "tree lucerne seedling ffc": "Tagasaste - Tree Lucerne",
    "Sweet basil": "Basil - Sweet (Classic)",
    "Cape Gooseberry": "Cape Gooseberry",
    "Pigeon pea": "Pigeon Pea",
    "Pigeon Pea": "Pigeon Pea",
    "oregano": "Oregano",
    "Ice ream bean": "Ice Cream Bean",
    "Ice-cream bean": "Ice Cream Bean",
    "Ice cream bean": "Ice Cream Bean",
    "Tamarin": "Tamarind",
    "Tomato": "Tomato (Marmande)",
    "Tomatoes": "Tomato (Marmande)",
    "tomatoes": "Tomato (Marmande)",
    # R5 Kolala Day native species
    "Rough barked apple": "Rough Barked Apple",
    "Prickly-leaved paperbark or prickly tea-tree": "Prickly-leaved Paperbark",
    "Black she-oak": None,  # Allocasuarina littoralis — no farmOS asset (dead)
    "Flax-leaved paperbark, snow in summer": None,  # Melaleuca linariifolia — dead
    "Prickly-leaved paperbark, prickly tea-tree": None,  # Melaleuca Nodosa — uninventoried
    "White Feather, Honey Myrtle": None,  # Melaleuca decora — uninventoried
    # R5 early plantings (text dates)
    "Apple tree seedling": "Apple",
    "Avocado tree seedling": "Avocado",
}

# Non-plant rows to skip in renovation sheets
SKIP_ROWS = {
    "LIME", "lime", "LIME g/m2", "lime g/m2",
    "GREEN M: VETX", "GREEN M vetch, radish, oat",
    "green manure", "GREEN MANURE", "summer green manure", "SUMMER GREEN MANURE",
    "automn green manure", "WINTER GREEN MANURE",
    "placenta", "PLACENTA", "EMERGENT PLACENTA",
}

# ── Spreadsheet tab registry ────────────────────────────────────────────

R2_FILE = "2025.SPRING.P2R2.0-46.INVENTORYandRENOVATION (2).xlsx"
R3_FILE = "P2R3.0-63.2025.SPRING.INVENTORYandRENOVATION (2).xlsx"
R4_SPRING_FILE = "P2R4.2025.SPRING.INVENTORY.xlsx"
R5_FILE = "P2R5.JAN2026.REGISTRATION.xlsx"

TAB_REGISTRY = {
    # R1 tabs (in the R2 file)
    (R2_FILE, "R1.0-3.2025 spring renovation"):   ("r1", "P2R1.0-3"),
    (R2_FILE, "R1.3-9.2025 spring renovation"):   ("r1", "P2R1.3-9"),
    (R2_FILE, "R1.9-13.2025 spring renovation"):  ("r1", "P2R1.9-13"),
    (R2_FILE, "R1.13-22.2025 spring renovation"): ("r1", "P2R1.13-22"),
    # R2 tabs
    (R2_FILE, "R2.0-3"):                              ("r2", "P2R2.0-3"),
    (R2_FILE, "R2.3-7"):                              ("r2", "P2R2.3-7"),
    (R2_FILE, "R2.7-16"):                             ("r2", "P2R2.7-16"),
    (R2_FILE, "R2.16-23"):                            ("r2", "P2R2.16-23"),
    (R2_FILE, "R2.23-26"):                            ("r2", "P2R2.23-26"),
    (R2_FILE, "R2.28-37.2025 Spring renovation"):     ("r2", "P2R2.28-37"),
    (R2_FILE, "R2.37-46.2025 Spring renovation"):     ("r2", "P2R2.37-46"),
    # R3 tabs
    (R3_FILE, "P2R3.0-3.2025SpringRenovation"):       ("r3", "P2R3.0-3"),
    (R3_FILE, "P2R3.3-9.2025SpringRenovatio "):       ("r3", "P2R3.3-9"),
    (R3_FILE, "P2R3.9-14.2025SpringRenovation"):      ("r3", "P2R3.9-14"),
    (R3_FILE, "P2R3.15-21.2025SpringRenovation"):     ("r3", "P2R3.15-21"),
    (R3_FILE, "P2R3.26-37.SpringRenovation"):         ("r3", "P2R3.26-37"),
    (R3_FILE, "P2R3.41-63.SpringRenovation"):         ("r3_shifted", "P2R3.41-63"),
    # R4 spring planting tabs (registration format — different from renovation)
    (R4_SPRING_FILE, "P2R4.0-1.2025spring.PlantnSeed"):  ("r4_spring", "P2R4.0-1"),
    (R4_SPRING_FILE, "P2R4.6-13.2025spring.PlantnS"):    ("r4_spring_g", "P2R4.6-13"),
    (R4_SPRING_FILE, "P2R4.20-27.2025spring.PlantnS"):   ("r4_spring_shifted", "P2R4.20-27"),
    (R4_SPRING_FILE, "P2R4.35-44.2025spring.PlantnS"):   ("r4_spring_g", "P2R4.35-44"),
    # R5 historical planting entries (in the main registration file)
    (R5_FILE, "P2R5.0-8.JANV.2026"):       ("r5_historical", "P2R5.0-8"),
    (R5_FILE, "P2R5.14-22.JANV.2026"):     ("r5_historical", "P2R5.14-22"),
    (R5_FILE, "P2R5.29-38.JANV.2026"):     ("r5_historical", "P2R5.29-38"),
}


# ── Helpers ─────────────────────────────────────────────────────────────

def parse_date_cell(val):
    """Parse a date cell value (datetime object or string) to a datetime."""
    if val is None:
        return None
    if isinstance(val, datetime):
        dt = val
        # Fix year typos (2015 → 2025)
        if dt.year < 2020:
            dt = dt.replace(year=dt.year + 10)
        return dt
    s = str(val).strip()
    if not s:
        return None
    # Fix double-slash typos: "12/11//2025"
    s = s.replace("//", "/")
    # Try DD/MM/YYYY
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    # Try "25/10 to 13/11/2025" — extract last date
    m = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})\s*$', s)
    if m:
        try:
            return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    return None


def to_timestamp(dt):
    """Convert datetime to Unix timestamp with AEST timezone."""
    if dt is None:
        return int(datetime(2025, 4, 1, tzinfo=AEST).timestamp())
    return int(dt.replace(tzinfo=AEST).timestamp())


def parse_count_cell(val):
    """Parse an inventory/count cell. Returns (count, notes)."""
    if val is None:
        return None, ""
    if isinstance(val, (int, float)):
        n = float(val)
        count = int(n) if n == int(n) else n
        # Sanity check: counts > 500 are likely year values (e.g. 2025)
        if count > 500:
            return None, f"suspicious count {count}"
        return count, ""
    s = str(val).strip()
    if not s or s in ("#VALUE!",):
        return None, s
    s_upper = s.upper()
    if "DEAD" in s_upper or s_upper in ("REMOVED", "REMOVED "):
        return 0, s
    # Extract leading number: "3 (details)" → 3
    m = re.match(r'^[>]?(\d+\.?\d*)\s*(.*)', s)
    if m:
        n = float(m.group(1))
        count = int(n) if n == int(n) else n
        # Sanity check: counts > 500 are likely year values
        if count > 500:
            return None, f"suspicious count {count}; {m.group(2).strip()}"
        notes = m.group(2).strip("() ")
        return count, notes
    return None, s


def is_skip_row(species_raw):
    """Check if a species name is a non-plant row to skip."""
    if not species_raw:
        return True
    name = species_raw.strip()
    if name in SKIP_SPECIES or name in SKIP_ROWS:
        return True
    name_lower = name.lower()
    if any(kw in name_lower for kw in (
        "green manure", "green m:", "summer green", "winter green",
        "automn green", "lime g/m", "placenta", "totals", "gm total",
        "1 lm=",
    )):
        return True
    return False


def normalize_historical_species(raw_name, farmos_names):
    """Normalize a species name from historical sheets."""
    if not raw_name or is_skip_row(raw_name):
        return None
    name = raw_name.strip().rstrip(".")
    # Remove newlines
    name = name.replace("\n", " ").replace("\r", " ")
    # Try historical overrides first (None = explicitly skip)
    if name in HISTORICAL_SPECIES_OVERRIDES:
        result = HISTORICAL_SPECIES_OVERRIDES[name]
        if result is None:
            return None  # Explicitly skipped (e.g., multi-species entries)
        return result
    # Fall through to the standard normalize_species
    return normalize_species(name, farmos_names)


# ── Spreadsheet parsers ─────────────────────────────────────────────────

def parse_renovation_tab(ws, fmt, historical_id):
    """Parse a renovation tab and return section data.

    fmt: 'r1', 'r2', 'r3', 'r3_shifted'
    """
    # Column offsets by format
    if fmt == "r3":
        col_prev = 8    # H
        col_inv = 9     # I
        col_reno = 10   # J
        col_notes = 11  # K
        col_extra = 14  # N (dates/origin)
        inv_date_cell = ws.cell(2, 9)    # I2
        reno_date_cell = ws.cell(2, 10)  # J2
    elif fmt == "r3_shifted":
        col_prev = 9    # I
        col_inv = 10    # J
        col_reno = 11   # K
        col_notes = 12  # L
        col_extra = 14  # N
        inv_date_cell = ws.cell(2, 10)   # J2
        reno_date_cell = ws.cell(2, 11)  # K2
    else:  # r1, r2
        col_prev = 9    # I
        col_inv = 10    # J
        col_reno = 11   # K
        col_notes = 12  # L
        col_extra = 14  # N
        inv_date_cell = ws.cell(2, 10)   # J2
        reno_date_cell = ws.cell(2, 11)  # K2

    inventory_date = parse_date_cell(inv_date_cell.value)
    renovation_date = parse_date_cell(reno_date_cell.value)

    # Parse species rows (start at row 5 or 6)
    plants = []
    current_strata = None

    for row in range(5, ws.max_row + 1):
        # Check strata in column A
        strata_val = ws.cell(row, 1).value
        if strata_val:
            s = str(strata_val).strip().lower()
            if "emergent" in s:
                current_strata = "emergent"
            elif "high" in s:
                current_strata = "high"
            elif "medium" in s:
                current_strata = "medium"
            elif "low" in s:
                current_strata = "low"

        # Species in column C
        species_raw = ws.cell(row, 3).value
        if not species_raw or is_skip_row(str(species_raw)):
            continue
        species_raw = str(species_raw).strip()

        prev_count, prev_notes = parse_count_cell(ws.cell(row, col_prev).value)
        inv_count, inv_notes = parse_count_cell(ws.cell(row, col_inv).value)
        reno_count, reno_notes = parse_count_cell(ws.cell(row, col_reno).value)

        # Extra notes from notes column and extra column
        notes_val = ws.cell(row, col_notes).value
        extra_val = ws.cell(row, col_extra).value
        extra_notes = ""
        if notes_val:
            extra_notes += str(notes_val).strip()
        if extra_val:
            if extra_notes:
                extra_notes += "; "
            extra_notes += str(extra_val).strip()
        # Also check column O (15) for origin
        origin_val = ws.cell(row, 15).value
        if origin_val:
            if extra_notes:
                extra_notes += "; "
            extra_notes += str(origin_val).strip()

        plants.append({
            "species_raw": species_raw,
            "strata": current_strata,
            "initial_count": prev_count,
            "inventory_count": inv_count,
            "inventory_notes": inv_notes,
            "renovation_count": reno_count,
            "renovation_notes": reno_notes,
            "extra_notes": extra_notes,
        })

    return {
        "historical_id": historical_id,
        "inventory_date": inventory_date,
        "renovation_date": renovation_date,
        "plants": plants,
    }


def parse_r4_spring_tab(ws, fmt, historical_id):
    """Parse an R4 spring planting tab (registration format).

    fmt: 'r4_spring'        — qty in F, date in H (tab 0-1)
         'r4_spring_g'      — plant qty in G, seed qty in F, date in H (tabs 6-13, 35-44)
         'r4_spring_shifted' — mixed F/G, date in J (tab 20-27)
    """
    # Determine column positions
    if fmt == "r4_spring_shifted":
        col_date = 10     # J
        col_origin = 11   # K
    else:
        col_date = 8      # H
        col_origin = 9    # I

    plants = []

    for row in range(5, ws.max_row + 1):
        species_raw = ws.cell(row, 1).value  # Col A
        if not species_raw or is_skip_row(str(species_raw)):
            continue
        species_raw = str(species_raw).strip()
        if not species_raw or species_raw.lower() in ("rjukn", "rjunk"):
            continue

        ps_val = ws.cell(row, 5).value  # Col E: P/S/tubercule
        ps_str = str(ps_val).strip().upper() if ps_val else ""

        # Skip seed entries (S) — we only want plants
        if ps_str == "S":
            continue

        # Get quantity — try both F and G, prefer the numeric one
        f_val = ws.cell(row, 6).value  # Col F
        g_val = ws.cell(row, 7).value  # Col G

        count = None
        if fmt == "r4_spring":
            # Tab 0-1: qty always in F
            count, _ = parse_count_cell(f_val)
        elif fmt == "r4_spring_g":
            # Tabs 6-13, 35-44: plant qty in G
            count, _ = parse_count_cell(g_val)
            if count is None:
                count, _ = parse_count_cell(f_val)
        elif fmt == "r4_spring_shifted":
            # Tab 20-27: mixed — try F first, then G
            count, _ = parse_count_cell(f_val)
            if count is None:
                count, _ = parse_count_cell(g_val)

        # Get planting date
        planting_date = parse_date_cell(ws.cell(row, col_date).value)

        # Get origin
        origin_val = ws.cell(row, col_origin).value
        origin = str(origin_val).strip() if origin_val else ""

        if count is not None and count > 0:
            plants.append({
                "species_raw": species_raw,
                "strata": None,
                "initial_count": count,
                "inventory_count": None,
                "inventory_notes": "",
                "renovation_count": None,
                "renovation_notes": "",
                "extra_notes": f"Origin: {origin}" if origin else "",
                "planting_date": planting_date,
            })

    return {
        "historical_id": historical_id,
        "inventory_date": None,
        "renovation_date": None,
        "plants": plants,
        "format": "r4_spring",
    }


def parse_r5_historical(ws, fmt, historical_id):
    """Parse R5 tabs for historical entries predating the main Dec 30 planting.

    Finds entries where Col H date is before December 2025 (Kolala Day,
    May-June 2025 early plantings).
    """
    plants = []
    dec_cutoff = datetime(2025, 12, 1)

    for row in range(5, ws.max_row + 1):
        species_raw = ws.cell(row, 1).value  # Col A
        if not species_raw:
            continue
        species_raw = str(species_raw).strip()
        if not species_raw or is_skip_row(species_raw):
            continue

        ps_val = ws.cell(row, 5).value  # Col E: P/S
        ps_str = str(ps_val).strip().upper() if ps_val else ""
        if ps_str == "S":
            continue

        date_val = ws.cell(row, 8).value  # Col H
        planting_date = None
        is_historical = False

        if isinstance(date_val, datetime):
            if date_val < dec_cutoff:
                planting_date = date_val
                is_historical = True
        elif isinstance(date_val, str):
            s = date_val.strip().lower()
            if "may" in s or "june" in s or "jun" in s:
                planting_date = datetime(2025, 5, 15)  # approximate
                is_historical = True

        if not is_historical:
            continue

        # Get quantity planted (Col F)
        qty_val = ws.cell(row, 6).value
        count, _ = parse_count_cell(qty_val)

        # Get inventory count (Col I) for notes
        inv_val = ws.cell(row, 9).value
        inv_count, _ = parse_count_cell(inv_val)

        # Get origin (Col J)
        origin_val = ws.cell(row, 10).value
        origin = str(origin_val).strip() if origin_val else ""

        if count is not None and count > 0:
            plants.append({
                "species_raw": species_raw,
                "strata": None,
                "initial_count": count,
                "inventory_count": None,
                "inventory_notes": "",
                "renovation_count": None,
                "renovation_notes": "",
                "extra_notes": f"Origin: {origin}" if origin else "",
                "planting_date": planting_date,
            })

    return {
        "historical_id": historical_id,
        "inventory_date": None,
        "renovation_date": None,
        "plants": plants,
        "format": "r5_historical",
    }


# ── farmOS importer ─────────────────────────────────────────────────────

def get_farmos_config():
    """Load farmOS configuration from environment variables."""
    load_dotenv()
    config = {
        "hostname": os.getenv("FARMOS_URL"),
        "username": os.getenv("FARMOS_USERNAME"),
        "password": os.getenv("FARMOS_PASSWORD"),
        "client_id": os.getenv("FARMOS_CLIENT_ID", "farm"),
        "scope": os.getenv("FARMOS_SCOPE", "farm_manager"),
    }
    missing = [k for k in ("hostname", "username", "password") if not config[k]]
    if missing:
        print(f"ERROR: Missing env vars: {', '.join('FARMOS_' + k.upper() for k in missing)}")
        sys.exit(1)
    return config


class HistoricalImporter:
    """Imports historical logs into farmOS for existing plant assets."""

    def __init__(self, config, dry_run=False):
        self.config = config
        self.dry_run = dry_run
        self.client = None
        self.plant_type_cache = {}
        self.section_cache = {}
        self.plant_asset_cache = {}  # (farmos_name, section_id) → uuid
        self.stats = {
            "planted_logs": 0,
            "inventory_logs": 0,
            "renovation_logs": 0,
            "skipped": 0,
            "unmatched": 0,
            "failed": 0,
        }

    def connect(self):
        print(f"\nConnecting to {self.config['hostname']}...")
        try:
            self.client = farmOS_client(
                hostname=self.config["hostname"],
                client_id=self.config["client_id"],
                scope=self.config["scope"],
            )
            self.client.authorize(
                username=self.config["username"],
                password=self.config["password"],
            )
            print("Connected successfully.")
            return True
        except Exception as e:
            print(f"Authentication failed: {e}")
            return False

    def fetch_by_name(self, api_path, name):
        session = self.client.session
        encoded = urllib.parse.quote(name)
        url = f"/api/{api_path}?filter[name]={encoded}&page[limit]=50"
        resp = session.http_request(url)
        if resp.status_code != 200:
            return []
        return resp.json().get("data", [])

    def get_section_uuid(self, section_id):
        if section_id in self.section_cache:
            return self.section_cache[section_id]
        assets = self.fetch_by_name("asset/land", section_id)
        if assets:
            uuid = assets[0]["id"]
            self.section_cache[section_id] = uuid
            return uuid
        return None

    def find_plant_asset(self, farmos_name, section_id, planted_label):
        """Find an existing plant asset by reconstructing its name."""
        cache_key = (farmos_name, section_id)
        if cache_key in self.plant_asset_cache:
            return self.plant_asset_cache[cache_key]

        asset_name = f"{planted_label} - {farmos_name} - {section_id}"
        assets = self.fetch_by_name("asset/plant", asset_name)
        if assets:
            uuid = assets[0]["id"]
            self.plant_asset_cache[cache_key] = uuid
            return uuid

        self.plant_asset_cache[cache_key] = None
        return None

    def log_exists(self, log_type, log_name):
        """Check if a log with this name already exists."""
        logs = self.fetch_by_name(f"log/{log_type}", log_name)
        return len(logs) > 0

    def create_quantity(self, plant_id, count, adjustment="reset"):
        """Create a quantity entity for inventory count tracking."""
        payload = {
            "data": {
                "type": "quantity--standard",
                "attributes": {
                    "value": {"decimal": str(count)},
                    "measure": "count",
                    "label": "plants",
                    "inventory_adjustment": adjustment,
                },
                "relationships": {
                    "units": {
                        "data": {
                            "type": "taxonomy_term--unit",
                            "id": PLANT_UNIT_UUID,
                        }
                    },
                    "inventory_asset": {
                        "data": {
                            "type": "asset--plant",
                            "id": plant_id,
                        }
                    },
                },
            }
        }
        try:
            session = self.client.session
            resp = session.http_request(
                "/api/quantity/standard",
                method="POST",
                options={"json": payload},
            )
            return resp.json().get("data", {}).get("id")
        except Exception as e:
            print(f"    ! Failed to create quantity: {e}")
            return None

    def create_log(self, log_type, plant_uuid, section_uuid, timestamp,
                   log_name, count=None, adjustment="reset", notes=""):
        """Create a log in farmOS."""
        log_data = {
            "attributes": {
                "name": log_name,
                "timestamp": str(timestamp),
                "status": "done",
                "is_movement": True,
            },
            "relationships": {
                "asset": {
                    "data": [{"type": "asset--plant", "id": plant_uuid}]
                },
                "location": {
                    "data": [{"type": "asset--land", "id": section_uuid}]
                },
            },
        }
        if notes:
            log_data["attributes"]["notes"] = {"value": notes, "format": "default"}

        if count is not None and count > 0:
            quantity_id = self.create_quantity(plant_uuid, count, adjustment)
            if quantity_id:
                log_data["relationships"]["quantity"] = {
                    "data": [{"type": "quantity--standard", "id": quantity_id}]
                }

        try:
            result = self.client.log.send(log_type, log_data)
            return result.get("data", {}).get("id")
        except Exception as e:
            print(f"    ! Failed to create {log_type} log: {e}")
            return None

    # ── Main import ─────────────────────────────────────────────

    def import_all(self, fieldsheets_dir, sections_path,
                   row_filter=None, section_filter=None):
        """Main import process."""
        print(f"\n{'='*60}")
        print("FIREFLY CORNER FARM — Historical Log Importer (Phase H1)")
        print(f"{'='*60}")

        if self.dry_run:
            print("\nDRY RUN — No changes will be made\n")

        # Load sections.json for first_planted dates
        with open(sections_path) as f:
            sections_data = json.load(f)

        sections_info = {}  # section_id → {first_planted, planted_label}
        for sid, sec in sections_data.get("sections", {}).items():
            fp = sec.get("first_planted", "")
            sections_info[sid] = {
                "first_planted": fp,
                "planted_label": format_planted_label(fp),
            }

        # Load farmos_names for species normalization
        farmos_names = load_farmos_names()
        print(f"Loaded {len(farmos_names)} farmos names for species matching")

        # Connect to farmOS (unless dry-run)
        if not self.dry_run:
            if not self.connect():
                return False

        # Parse all renovation tabs
        all_sections = []
        workbooks = {}

        for (filename, tab_name), (fmt, hist_id) in TAB_REGISTRY.items():
            filepath = Path(fieldsheets_dir) / filename
            if not filepath.exists():
                print(f"  Warning: {filepath} not found, skipping")
                continue

            if filename not in workbooks:
                workbooks[filename] = openpyxl.load_workbook(
                    str(filepath), data_only=True
                )

            wb = workbooks[filename]
            if tab_name not in wb.sheetnames:
                print(f"  Warning: tab '{tab_name}' not found in {filename}")
                continue

            ws = wb[tab_name]

            # Dispatch to appropriate parser
            if fmt.startswith("r4_spring"):
                section_data = parse_r4_spring_tab(ws, fmt, hist_id)
            elif fmt == "r5_historical":
                section_data = parse_r5_historical(ws, fmt, hist_id)
            else:
                section_data = parse_renovation_tab(ws, fmt, hist_id)

            # Map to current section(s)
            current = HISTORICAL_TO_CURRENT.get(hist_id)
            if current is None:
                continue  # no current section (P2R3.21-26)

            if isinstance(current, list):
                section_data["current_ids"] = current
            else:
                section_data["current_ids"] = [current]

            all_sections.append(section_data)

        # Close workbooks
        for wb in workbooks.values():
            wb.close()

        # Apply filters
        if row_filter:
            all_sections = [s for s in all_sections
                           if any(cid.startswith(row_filter) for cid in s["current_ids"])]
            print(f"Filtered to {len(all_sections)} sections matching {row_filter}")

        if section_filter:
            all_sections = [s for s in all_sections
                           if section_filter in s["current_ids"]]
            print(f"Filtered to {len(all_sections)} sections matching {section_filter}")

        # Count work
        total_plants = sum(len(s["plants"]) for s in all_sections)
        print(f"Processing {len(all_sections)} sections, {total_plants} species rows\n")

        # Process each section
        for section_data in all_sections:
            self.process_section(section_data, sections_info, farmos_names)

        # Summary
        print(f"\n{'='*60}")
        print("IMPORT SUMMARY")
        print(f"{'='*60}")
        print(f"  Planted logs:     {self.stats['planted_logs']}")
        print(f"  Inventory logs:   {self.stats['inventory_logs']}")
        print(f"  Renovation logs:  {self.stats['renovation_logs']}")
        print(f"  Skipped (exist):  {self.stats['skipped']}")
        print(f"  Unmatched:        {self.stats['unmatched']}")
        print(f"  Failed:           {self.stats['failed']}")

        if self.dry_run:
            print(f"\n  ** DRY RUN — run without --dry-run to apply changes **")
        else:
            print(f"\n  Import completed!")

        return True

    def process_section(self, section_data, sections_info, farmos_names):
        """Process one historical section, creating logs for each species."""
        hist_id = section_data["historical_id"]
        current_ids = section_data["current_ids"]
        inv_date = section_data["inventory_date"]
        reno_date = section_data["renovation_date"]

        for current_id in current_ids:
            info = sections_info.get(current_id)
            if not info:
                print(f"  ! {current_id}: not found in sections.json")
                continue

            planted_label = info["planted_label"]
            first_planted = info["first_planted"]

            # Parse initial planting date
            initial_date = parse_initial_date(first_planted)

            section_uuid = None
            if not self.dry_run:
                section_uuid = self.get_section_uuid(current_id)
                if not section_uuid:
                    print(f"  ! {current_id}: not found in farmOS")
                    continue

            active_plants = [p for p in section_data["plants"]
                            if not is_skip_row(p["species_raw"])]
            if not active_plants:
                continue

            print(f"  {hist_id} → {current_id}: {len(active_plants)} species")

            for plant in active_plants:
                farmos_name = normalize_historical_species(
                    plant["species_raw"], farmos_names
                )
                if not farmos_name:
                    continue

                # Find existing plant asset
                plant_uuid = None
                if not self.dry_run:
                    plant_uuid = self.find_plant_asset(
                        farmos_name, current_id, planted_label
                    )
                    if not plant_uuid:
                        print(f"    ? {farmos_name} not found in {current_id}")
                        self.stats["unmatched"] += 1
                        continue

                # 1. Initial planting log
                if plant["initial_count"] is not None and plant["initial_count"] > 0:
                    log_name = f"Planted {current_id} — {farmos_name}"
                    notes = plant.get("extra_notes", "")

                    # Use per-plant date if available (R4/R5), else section-level date
                    plant_date = plant.get("planting_date") or initial_date

                    if self.dry_run:
                        date_str = plant_date.strftime("%Y-%m-%d") if plant_date else "unknown"
                        print(f"    + PLANTED: {farmos_name} (count: {plant['initial_count']}, date: {date_str})")
                        self.stats["planted_logs"] += 1
                    else:
                        if self.log_exists("transplanting", log_name):
                            self.stats["skipped"] += 1
                        else:
                            ts = to_timestamp(plant_date)
                            log_id = self.create_log(
                                "transplanting", plant_uuid, section_uuid, ts,
                                log_name, count=plant["initial_count"],
                                adjustment="reset", notes=notes,
                            )
                            if log_id:
                                self.stats["planted_logs"] += 1
                            else:
                                self.stats["failed"] += 1

                # 2. Mid-season observation log
                if plant["inventory_count"] is not None and inv_date:
                    log_name = f"Inventory Oct 2025 {current_id} — {farmos_name}"
                    notes_parts = []
                    if plant["inventory_notes"]:
                        notes_parts.append(plant["inventory_notes"])
                    obs_notes = "; ".join(notes_parts)

                    if self.dry_run:
                        print(f"    ~ INVENTORY: {farmos_name} (count: {plant['inventory_count']})")
                        self.stats["inventory_logs"] += 1
                    else:
                        if self.log_exists("observation", log_name):
                            self.stats["skipped"] += 1
                        else:
                            ts = to_timestamp(inv_date)
                            log_id = self.create_log(
                                "observation", plant_uuid, section_uuid, ts,
                                log_name, count=plant["inventory_count"],
                                adjustment="reset", notes=obs_notes,
                            )
                            if log_id:
                                self.stats["inventory_logs"] += 1
                            else:
                                self.stats["failed"] += 1

                # 3. Renovation planting log
                if plant["renovation_count"] is not None and plant["renovation_count"] > 0:
                    log_name = f"Renovation {current_id} — {farmos_name}"
                    notes = plant.get("renovation_notes", "")
                    if plant.get("extra_notes"):
                        if notes:
                            notes += "; "
                        notes += plant["extra_notes"]

                    if self.dry_run:
                        print(f"    + RENOVATION: {farmos_name} (count: {plant['renovation_count']})")
                        self.stats["renovation_logs"] += 1
                    else:
                        if self.log_exists("transplanting", log_name):
                            self.stats["skipped"] += 1
                        else:
                            ts = to_timestamp(reno_date or inv_date)
                            log_id = self.create_log(
                                "transplanting", plant_uuid, section_uuid, ts,
                                log_name, count=plant["renovation_count"],
                                adjustment="increment", notes=notes,
                            )
                            if log_id:
                                self.stats["renovation_logs"] += 1
                            else:
                                self.stats["failed"] += 1


def format_planted_label(date_str):
    """Format first_planted date for plant asset name lookup."""
    if not date_str:
        return "SPRING 2025"
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        return dt.strftime("%-d %b %Y").upper()
    except ValueError:
        pass
    try:
        dt = datetime.strptime(date_str, "%B %Y")
        return dt.strftime("%b %Y").upper()
    except ValueError:
        pass
    return date_str.upper()


def parse_initial_date(first_planted_str):
    """Parse the first_planted date string from sections.json."""
    if not first_planted_str:
        return datetime(2025, 4, 1)
    try:
        return datetime.strptime(first_planted_str, "%Y-%m-%d")
    except ValueError:
        pass
    try:
        return datetime.strptime(first_planted_str, "%B %Y")
    except ValueError:
        pass
    return datetime(2025, 4, 1)


# ── CLI ─────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Import historical planting/observation data into farmOS"
    )
    parser.add_argument(
        "--data", default="site/src/data/sections.json",
        help="Path to sections.json",
    )
    parser.add_argument(
        "--fieldsheets", default="fieldsheets/",
        help="Path to fieldsheets directory",
    )
    parser.add_argument(
        "--row", default=None,
        help="Filter to row (e.g., P2R1, P2R2, P2R3)",
    )
    parser.add_argument(
        "--section", default=None,
        help="Filter to section (e.g., P2R2.0-3)",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Preview without making changes",
    )
    args = parser.parse_args()

    data_path = Path(args.data)
    if not data_path.exists():
        print(f"Error: {data_path} not found")
        sys.exit(1)

    config = get_farmos_config()
    importer = HistoricalImporter(config, dry_run=args.dry_run)
    success = importer.import_all(
        args.fieldsheets, str(data_path),
        row_filter=args.row, section_filter=args.section,
    )
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
